#!/usr/bin/env python3
"""
European Lobbying Data Tool

Fetches comprehensive lobbying data from multiple European registries:

1. EU (via LobbyFacts.eu):
   - Registration history and lobbying expenditure over time
   - High-level Commission meetings (Commissioners, Cabinet members, DGs)

2. France (via HATVP):
   - Registration and lobbying expenditure
   - Detailed lobbying activities with subjects
   - Which ministries/officials were targeted

3. Germany (via Bundestag Lobbyregister):
   - Registration and lobbying expenditure ranges
   - Employee information and Berlin office presence
   - Regulatory projects targeted

4. Ireland (via Lobbying.ie - manual CSV download):
   - Lobbying returns and activities
   - Public officials and bodies contacted
   - Note: Ireland does NOT require financial disclosure

5. UK (via GOV.UK Ministerial Transparency):
   - Ministerial meetings with external organisations
   - Published quarterly by each government department
   - Note: UK does NOT track lobbying expenditure

6. Austria (via Lobbying- und Interessenvertretungs-Register):
   - Lobbying firms, in-house lobbyists, interest groups
   - Names of individual lobbyists
   - Financial data only disclosed if >€100,000

7. Catalonia (via Registre de grups d'interès de Catalunya):
   - Interest groups registered with Generalitat and Parliament
   - Annual business volume and areas of interest
   - Covers regional and local government lobbying

8. Finland (via Avoimuusrekisteri):
   - Lobbying activities targeting Parliament and Ministries
   - Activity disclosures submitted twice yearly
   - Lobbying topics and targets contacted
   - Financial data available from July 2026

Usage:
    python eu_lobbying.py --company "Google"
    python eu_lobbying.py --company "Microsoft" --output ms_report.xlsx
    python eu_lobbying.py --company "Google" --skip-uk  # Faster, skip UK search
    python eu_lobbying.py --ie-help  # Instructions for Irish data download
"""

import argparse
import csv
import io
import re
import sys
import zipfile
from datetime import datetime
from pathlib import Path

import requests
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# URLs
LOBBYFACTS_CSV_URL = "https://www.lobbyfacts.eu/csv_export/{id}"
LOBBYFACTS_MEETINGS_URL = "https://www.lobbyfacts.eu/csv_export_meetings/{id}"
HATVP_CSV_URL = "https://www.hatvp.fr/agora/opendata/csv/Vues_Separees_CSV.zip"
EU_REGISTER_URL = "https://transparency-register.europa.eu/odplastorganisationxml_en"
DE_LOBBYREGISTER_SEARCH_URL = "https://www.lobbyregister.bundestag.de/sucheDetailJson"
DE_LOBBYREGISTER_DETAIL_URL = "https://www.lobbyregister.bundestag.de/sucheJson/{reg_num}/{entry_id}"
# Ireland - Note: The Irish lobbying.ie API requires browser-based authentication
# Users must manually download CSV from https://www.lobbying.ie/app/home/search
IE_LOBBY_SEARCH_URL = "https://www.lobbying.ie/app/home/search"
IE_LOBBY_ORG_URL = "https://www.lobbying.ie/organisation/{org_id}"

CACHE_DIR = Path.home() / ".cache" / "eu_lobbying"


def get_cache_dir():
    CACHE_DIR.mkdir(parents=True, exist_ok=True)
    return CACHE_DIR


def search_eu_register(search_term: str) -> list:
    """Search EU Transparency Register for matching organisations."""
    from lxml import etree
    
    cache_file = get_cache_dir() / "eu_transparency_register.xml"
    
    if not cache_file.exists():
        print("Downloading EU Transparency Register (first time, ~100MB)...")
        response = requests.get(EU_REGISTER_URL, stream=True, timeout=120)
        response.raise_for_status()
        with open(cache_file, 'wb') as f:
            for chunk in response.iter_content(chunk_size=8192):
                f.write(chunk)
        print("Download complete!")
    
    parser = etree.XMLParser(recover=True, encoding='utf-8')
    tree = etree.parse(str(cache_file), parser)
    root = tree.getroot()
    
    results = []
    search_lower = search_term.lower()
    
    for elem in root.iter("interestRepresentative"):
        name_elem = elem.find("name/originalName")
        id_elem = elem.find("identificationCode")
        acronym_elem = elem.find("acronym")
        
        name = name_elem.text if name_elem is not None and name_elem.text else ""
        org_id = id_elem.text if id_elem is not None and id_elem.text else ""
        acronym = acronym_elem.text if acronym_elem is not None and acronym_elem.text else ""
        
        if search_lower in name.lower() or search_lower in acronym.lower():
            results.append({"name": name, "id": org_id, "acronym": acronym})
    
    return results


def search_france_register(search_term: str) -> list:
    """Search French HATVP register for matching organisations."""
    cache_file = get_cache_dir() / "hatvp_data.zip"
    extract_dir = get_cache_dir() / "hatvp"
    
    # Download if not cached or older than 24 hours
    if not cache_file.exists() or (datetime.now().timestamp() - cache_file.stat().st_mtime) > 86400:
        print("Downloading French HATVP register...")
        response = requests.get(HATVP_CSV_URL, timeout=120)
        response.raise_for_status()
        with open(cache_file, 'wb') as f:
            f.write(response.content)
        
        # Extract
        with zipfile.ZipFile(cache_file, 'r') as zf:
            zf.extractall(extract_dir)
        print("Download complete!")
    elif not extract_dir.exists():
        with zipfile.ZipFile(cache_file, 'r') as zf:
            zf.extractall(extract_dir)
    
    # Search
    results = []
    search_lower = search_term.lower()
    info_file = extract_dir / "Vues_Separees" / "1_informations_generales.csv"
    
    with open(info_file, 'r', encoding='utf-8') as f:
        reader = csv.DictReader(f, delimiter=';')
        for row in reader:
            name = row.get('denomination', '')
            if search_lower in name.lower():
                results.append({
                    "name": name,
                    "id": row.get('representants_id'),
                    "siren": row.get('identifiant_national', ''),
                    "city": row.get('ville', ''),
                    "category": row.get('label_categorie_organisation', '')
                })
    
    return results


def fetch_eu_data(org_id: str) -> dict:
    """Fetch organisation data from LobbyFacts."""
    print(f"Fetching EU LobbyFacts data for ID: {org_id}")
    
    # Fetch registrations
    response = requests.get(LOBBYFACTS_CSV_URL.format(id=org_id), timeout=60)
    response.raise_for_status()
    registrations = list(csv.DictReader(io.StringIO(response.text)))
    
    # Fetch meetings
    response = requests.get(LOBBYFACTS_MEETINGS_URL.format(id=org_id), timeout=60)
    response.raise_for_status()
    meetings = list(csv.DictReader(io.StringIO(response.text.lstrip('\ufeff'))))
    
    print(f"  Found {len(registrations)} registration snapshots, {len(meetings)} Commission meetings")
    
    return {
        "registrations": registrations, 
        "meetings": meetings, 
        "org_id": org_id,
        "data_coverage": "2012-present"
    }


def fetch_france_data(org_id: str) -> dict:
    """Fetch organisation data from French HATVP."""
    print(f"Fetching French HATVP data for ID: {org_id}")
    
    extract_dir = get_cache_dir() / "hatvp" / "Vues_Separees"
    
    # Get general info
    info = {}
    with open(extract_dir / "1_informations_generales.csv", 'r', encoding='utf-8') as f:
        reader = csv.DictReader(f, delimiter=';')
        for row in reader:
            if row.get('representants_id') == org_id:
                info = row
                break
    
    # Get exercises (yearly reports)
    exercises = []
    with open(extract_dir / "15_exercices.csv", 'r', encoding='utf-8') as f:
        reader = csv.DictReader(f, delimiter=';')
        for row in reader:
            if row.get('representants_id') == org_id:
                exercises.append(row)
    
    exercise_ids = [e['exercices_id'] for e in exercises]
    
    # Get activities
    activities = []
    with open(extract_dir / "8_objets_activites.csv", 'r', encoding='utf-8') as f:
        reader = csv.DictReader(f, delimiter=';')
        for row in reader:
            if row.get('exercices_id') in exercise_ids:
                activities.append(row)
    
    activity_ids = set(a['activite_id'] for a in activities)
    
    # Get officials targeted
    officials = []
    with open(extract_dir / "13_ministeres_aai_api.csv", 'r', encoding='utf-8') as f:
        reader = csv.DictReader(f, delimiter=';')
        for row in reader:
            if row.get('action_representation_interet_id') in activity_ids:
                officials.append(row)
    
    # Get decisions targeted
    decisions = []
    with open(extract_dir / "12_decisions_concernees.csv", 'r', encoding='utf-8') as f:
        reader = csv.DictReader(f, delimiter=';')
        for row in reader:
            if row.get('action_representation_interet_id') in activity_ids:
                decisions.append(row)
    
    print(f"  Found {len(exercises)} years of data, {len(activities)} lobbying activities")
    
    return {
        "info": info,
        "exercises": exercises,
        "activities": activities,
        "officials": officials,
        "decisions": decisions,
        "org_id": org_id,
        "data_coverage": "2017-present"
    }


def search_germany_register(search_term: str) -> list:
    """Search German Bundestag Lobbyregister for matching organisations."""
    print(f"Searching German Lobbyregister for '{search_term}'...")
    
    url = f"{DE_LOBBYREGISTER_SEARCH_URL}?q={search_term}&sort=RELEVANCE_DESC"
    response = requests.get(url, timeout=60)
    response.raise_for_status()
    data = response.json()
    
    results = []
    for r in data.get("results", []):
        identity = r.get("lobbyistIdentity", {})
        results.append({
            "name": identity.get("name", "Unknown"),
            "register_number": r.get("registerNumber", ""),
            "entry_id": r.get("registerEntryDetails", {}).get("registerEntryId", ""),
            "city": identity.get("address", {}).get("city", ""),
            "active": r.get("accountDetails", {}).get("activeLobbyist", False)
        })
    
    return results


def fetch_germany_data(register_number: str, entry_id: str = None) -> dict:
    """Fetch organisation data from German Bundestag Lobbyregister."""
    print(f"Fetching German Lobbyregister data for: {register_number}")
    
    # If no entry_id, search for it first
    if not entry_id:
        url = f"{DE_LOBBYREGISTER_SEARCH_URL}?q={register_number}&sort=RELEVANCE_DESC"
        response = requests.get(url, timeout=60)
        response.raise_for_status()
        data = response.json()
        if data.get("results"):
            entry_id = data["results"][0].get("registerEntryDetails", {}).get("registerEntryId")
    
    if not entry_id:
        raise ValueError(f"Could not find entry for {register_number}")
    
    # Fetch full details
    url = DE_LOBBYREGISTER_DETAIL_URL.format(reg_num=register_number, entry_id=entry_id)
    response = requests.get(url, timeout=60)
    response.raise_for_status()
    data = response.json()
    
    # Extract key info
    identity = data.get("lobbyistIdentity", {})
    finances = data.get("financialExpenses", {})
    employees = data.get("employeesInvolvedInLobbying", {})
    activities = data.get("activitiesAndInterests", {})
    reg_projects = data.get("regulatoryProjects", {})
    
    # Get version history for historical data
    versions = data.get("accountDetails", {}).get("registerEntryVersions", [])
    
    result = {
        "register_number": register_number,
        "name": identity.get("name", ""),
        "legal_form": identity.get("legalForm", {}).get("de", ""),
        "city": identity.get("address", {}).get("city", ""),
        "country": identity.get("address", {}).get("country", {}).get("de", ""),
        "berlin_office": identity.get("capitalCityRepresentationPresent", False),
        "berlin_address": identity.get("capitalCityRepresentation", {}).get("address", {}).get("street", ""),
        "contact_email": identity.get("contactDetails", {}).get("emails", [{}])[0].get("email", "") if identity.get("contactDetails", {}).get("emails") else "",
        
        # Financial
        "fiscal_year_start": finances.get("relatedFiscalYearStart", ""),
        "fiscal_year_end": finances.get("relatedFiscalYearEnd", ""),
        "expenses_min": finances.get("financialExpensesEuro", {}).get("from", 0),
        "expenses_max": finances.get("financialExpensesEuro", {}).get("to", 0),
        
        # Employees
        "employee_fte": employees.get("employeeFTE", 0),
        
        # Activities
        "activity_type": activities.get("activity", {}).get("de", ""),
        "fields_of_interest": [f.get("de", f.get("en", "")) for f in activities.get("fieldsOfInterest", [])],
        
        # Regulatory projects
        "regulatory_projects_count": reg_projects.get("regulatoryProjectsCount", 0),
        "regulatory_projects": [],
        
        # Version history
        "version_count": len(versions),
        "first_publication": data.get("accountDetails", {}).get("firstPublicationDate", ""),
        "last_update": data.get("accountDetails", {}).get("lastUpdateDate", ""),
        
        # Data coverage
        "data_coverage": "2022-present",
        
        # Raw data for Excel
        "_raw": data
    }
    
    # Extract regulatory projects
    for proj in reg_projects.get("regulatoryProjects", []):
        result["regulatory_projects"].append({
            "title": proj.get("title", ""),
            "number": proj.get("regulatoryProjectNumber", ""),
            "printed_matters": [
                {
                    "title": pm.get("title", ""),
                    "number": pm.get("printingNumber", ""),
                    "issuer": pm.get("issuer", ""),
                    "url": pm.get("documentUrl", ""),
                    "ministry": pm.get("leadingMinistries", [{}])[0].get("shortTitle", "") if pm.get("leadingMinistries") else ""
                }
                for pm in proj.get("printedMatters", [])
            ]
        })
    
    print(f"  Found: {result['name']}")
    print(f"  Lobbying expenses: €{result['expenses_min']:,} - €{result['expenses_max']:,}")
    print(f"  Employee FTE: {result['employee_fte']}")
    print(f"  Regulatory projects: {result['regulatory_projects_count']}")
    
    return result


def load_ireland_csv(csv_path: str, search_term: str = None) -> dict:
    """
    Load Irish lobbying data from manually downloaded CSV.
    
    The Irish lobbying.ie website doesn't provide a public API.
    Users must manually download CSV from: https://www.lobbying.ie/app/home/search
    
    Steps:
    1. Go to https://www.lobbying.ie/app/home/search
    2. Search for your organisation (e.g., "Google")
    3. Click "Export to CSV" button
    4. Pass the downloaded CSV to this tool with --ie-csv
    """
    print(f"Loading Irish lobbying data from: {csv_path}")
    
    returns = []
    organisations = {}
    
    with open(csv_path, 'r', encoding='utf-8-sig') as f:
        reader = csv.DictReader(f)
        for row in reader:
            # Filter by search term if provided
            if search_term:
                lobbyist_name = row.get('Lobbyist', '').lower()
                client = row.get('Client', '').lower()
                if search_term.lower() not in lobbyist_name and search_term.lower() not in client:
                    continue
            
            lobbyist = row.get('Lobbyist', '')
            if lobbyist not in organisations:
                organisations[lobbyist] = {
                    "name": lobbyist,
                    "trading_name": row.get('Trading Name', ''),
                    "returns_count": 0
                }
            organisations[lobbyist]["returns_count"] += 1
            
            returns.append({
                "lobbyist": lobbyist,
                "trading_name": row.get('Trading Name', ''),
                "client": row.get('Client', ''),
                "subject_matter": row.get('Subject Matter', ''),
                "subject_matter_area": row.get('Subject Matter Area', ''),
                "subject_matter_details": row.get('Subject Matter Details', ''),
                "intended_result": row.get('Intended Results', ''),
                "public_body": row.get('Public Body', ''),
                "dpo_name": row.get('Designated Public Official', ''),
                "dpo_title": row.get('Job Title', ''),
                "return_period": row.get('Return Period', ''),
                "submitted_date": row.get('Submitted Date', ''),
                "nil_return": row.get('Nil Return', '') == 'True',
            })
    
    if not returns:
        print("  No matching records found in CSV")
        return None
    
    # Get primary organisation
    primary_org = max(organisations.values(), key=lambda x: x['returns_count'])
    
    result = {
        "name": primary_org["name"],
        "trading_name": primary_org.get("trading_name", ""),
        "returns": returns,
        "returns_count": len(returns),
        "organisations": list(organisations.values()),
        # Ireland doesn't track financial data
        "note": "Ireland's Regulation of Lobbying Act 2015 does not require disclosure of lobbying expenditure"
    }
    
    print(f"  Found: {result['name']}")
    print(f"  Returns: {result['returns_count']}")
    print(f"  Note: Ireland does not track lobbying expenditure")
    
    return result


def print_ireland_instructions():
    """Print instructions for manually downloading Irish lobbying data."""
    print("""
    ═══════════════════════════════════════════════════════════════════════════
    IRELAND LOBBYING DATA - MANUAL DOWNLOAD REQUIRED
    ═══════════════════════════════════════════════════════════════════════════
    
    The Irish lobbying register (lobbying.ie) doesn't provide a public API.
    To include Irish data, please manually download the CSV:
    
    1. Go to: https://www.lobbying.ie/app/home/search
    2. Enter your search term (e.g., "Google") in the search box
    3. Click the "Search" button
    4. Click "Export to CSV" (download icon) to download results
    5. Run this tool with: --ie-csv path/to/downloaded.csv
    
    The Irish register includes:
    - Lobbyist name and trading name
    - Subject matter and policy area
    - Public body and official contacted
    - Return periods (4-month intervals)
    
    Note: Unlike EU/France/Germany, Ireland does NOT track lobbying expenditure.
    ═══════════════════════════════════════════════════════════════════════════
    """)


# UK GOV.UK Search API and Content API for ministerial transparency data
UK_GOVUK_SEARCH_API = "https://www.gov.uk/api/search.json"
UK_GOVUK_CONTENT_API = "https://www.gov.uk/api/content"
UK_MEETINGS_CACHE_DIR = Path.home() / ".cache" / "eu_lobbying" / "uk_meetings"
UK_PUBLICATIONS_CACHE = Path.home() / ".cache" / "eu_lobbying" / "uk_publications_index.json"


def discover_uk_transparency_publications(max_results: int = 500) -> list:
    """
    Dynamically discover all UK ministerial meetings transparency publications
    using the GOV.UK Search API.
    
    Returns list of publication paths that can be fetched via Content API.
    Results are cached for 24 hours.
    """
    UK_MEETINGS_CACHE_DIR.mkdir(parents=True, exist_ok=True)
    
    # Check cache first
    if UK_PUBLICATIONS_CACHE.exists():
        cache_age = datetime.now().timestamp() - UK_PUBLICATIONS_CACHE.stat().st_mtime
        if cache_age < 86400:  # 24 hours
            try:
                with open(UK_PUBLICATIONS_CACHE, 'r') as f:
                    cached = json.load(f)
                    if cached.get('publications'):
                        return cached['publications']
            except:
                pass
    
    print("  Discovering UK transparency publications via GOV.UK Search API...")
    
    publications = []
    start = 0
    batch_size = 100  # GOV.UK API max is 1500, but we paginate
    
    while start < max_results:
        params = {
            "filter_format": "transparency",
            "q": "ministerial meetings",
            "count": batch_size,
            "start": start,
            "fields": "title,link,organisations,public_timestamp"
        }
        
        try:
            response = requests.get(UK_GOVUK_SEARCH_API, params=params, timeout=60)
            response.raise_for_status()
            data = response.json()
            
            results = data.get("results", [])
            if not results:
                break
                
            for r in results:
                title = r.get("title", "").lower()
                # Filter for meetings publications (not travel, gifts, hospitality only)
                if "meeting" in title:
                    org = ""
                    orgs = r.get("organisations", [])
                    if orgs and isinstance(orgs, list) and len(orgs) > 0:
                        org = orgs[0].get("title", "") if isinstance(orgs[0], dict) else ""
                    
                    publications.append({
                        "title": r.get("title", ""),
                        "link": r.get("link", ""),
                        "organisation": org,
                        "date": r.get("public_timestamp", "")
                    })
            
            start += batch_size
            
            # If we got fewer results than requested, we've reached the end
            if len(results) < batch_size:
                break
                
        except Exception as e:
            print(f"  Warning: Error fetching publications page {start}: {e}")
            break
    
    # Cache the results
    try:
        with open(UK_PUBLICATIONS_CACHE, 'w') as f:
            json.dump({"publications": publications, "timestamp": datetime.now().isoformat()}, f)
    except:
        pass
    
    print(f"  Found {len(publications)} ministerial meetings publications")
    return publications


def get_csv_urls_from_publication(publication_path: str) -> list:
    """
    Fetch a publication from GOV.UK Content API and extract CSV URLs.
    
    Returns list of tuples: (department, csv_url)
    """
    csv_urls = []
    
    try:
        url = f"{UK_GOVUK_CONTENT_API}{publication_path}"
        response = requests.get(url, timeout=30)
        response.raise_for_status()
        data = response.json()
        
        # Get organisation name
        orgs = data.get("links", {}).get("organisations", [])
        org_name = orgs[0].get("title", "Unknown") if orgs else "Unknown"
        
        # Extract CSV URLs from the documents HTML
        details = data.get("details", {})
        documents = details.get("documents", [])
        
        for doc in documents:
            # Documents contain HTML with attachment links
            # Extract URLs ending in .csv from href attributes
            urls = re.findall(r'href="(https://assets\.publishing\.service\.gov\.uk/[^"]+\.csv)"', doc)
            for csv_url in urls:
                # Only include meetings CSVs, not travel/gifts
                csv_lower = csv_url.lower()
                if 'meeting' in csv_lower and 'travel' not in csv_lower:
                    csv_urls.append((org_name, csv_url))
        
    except Exception as e:
        pass  # Skip problematic publications silently
    
    return csv_urls


def download_uk_meetings_csv(url: str, cache_key: str) -> str:
    """Download a UK ministerial meetings CSV, with caching."""
    UK_MEETINGS_CACHE_DIR.mkdir(parents=True, exist_ok=True)
    cache_file = UK_MEETINGS_CACHE_DIR / f"{cache_key}.csv"
    
    # Use cached version if less than 24 hours old
    if cache_file.exists() and (datetime.now().timestamp() - cache_file.stat().st_mtime) < 86400:
        return str(cache_file)
    
    response = requests.get(url, timeout=60)
    response.raise_for_status()
    
    with open(cache_file, 'wb') as f:
        f.write(response.content)
    
    return str(cache_file)


def parse_uk_csv_for_matches(csv_path: str, search_lower: str, dept: str) -> list:
    """Parse a UK ministerial/senior officials meetings CSV and return matching rows."""
    matches = []
    try:
        with open(csv_path, 'r', encoding='utf-8-sig', errors='replace') as f:
            reader = csv.DictReader(f)
            for row in reader:
                # Search in organisation name and purpose - handle various column names (case-insensitive)
                # Create lowercase key lookup
                row_lower = {k.lower(): v for k, v in row.items()}
                
                org_name = (row_lower.get('name of individual or organisation') or 
                           row_lower.get('name of organisation or individual') or 
                           row_lower.get('organisation') or 
                           row_lower.get('name of organisation') or '')
                purpose = (row_lower.get('purpose of meeting') or 
                          row_lower.get('purpose') or '')
                
                # Get official name - could be Minister or Senior Official
                official = (row_lower.get('minister') or 
                           row_lower.get("senior official's name") or
                           row_lower.get('senior official') or
                           row_lower.get('name') or '')
                
                date = row_lower.get('date', '')
                
                if search_lower in org_name.lower() or search_lower in purpose.lower():
                    matches.append({
                        "minister": official,
                        "date": date,
                        "organisation": org_name,
                        "purpose": purpose,
                        "department": dept,
                        "source": csv_path.split('/')[-1] if '/' in csv_path else csv_path.split('\\')[-1]
                    })
    except Exception as e:
        pass  # Skip problematic files silently
    return matches


UK_SENIOR_OFFICIALS_CACHE = Path.home() / ".cache" / "eu_lobbying" / "uk_senior_officials_index.json"


def discover_uk_senior_officials_publications(max_results: int = 300) -> list:
    """
    Dynamically discover UK senior officials meetings transparency publications
    using the GOV.UK Search API.
    
    Returns list of publication paths that can be fetched via Content API.
    Results are cached for 24 hours.
    """
    UK_MEETINGS_CACHE_DIR.mkdir(parents=True, exist_ok=True)
    
    # Check cache first
    if UK_SENIOR_OFFICIALS_CACHE.exists():
        cache_age = datetime.now().timestamp() - UK_SENIOR_OFFICIALS_CACHE.stat().st_mtime
        if cache_age < 86400:  # 24 hours
            try:
                with open(UK_SENIOR_OFFICIALS_CACHE, 'r') as f:
                    cached = json.load(f)
                    if cached.get('publications'):
                        return cached['publications']
            except:
                pass
    
    print("  Discovering UK senior officials publications via GOV.UK Search API...")
    
    publications = []
    start = 0
    batch_size = 100
    
    while start < max_results:
        params = {
            "filter_format": "transparency",
            "q": "senior officials meetings",
            "count": batch_size,
            "start": start,
            "fields": "title,link,organisations,public_timestamp"
        }
        
        try:
            response = requests.get(UK_GOVUK_SEARCH_API, params=params, timeout=60)
            response.raise_for_status()
            data = response.json()
            
            results = data.get("results", [])
            if not results:
                break
                
            for r in results:
                title = r.get("title", "").lower()
                # Filter for meetings publications (not expenses, hospitality only)
                if "meeting" in title and "senior official" in title:
                    org = ""
                    orgs = r.get("organisations", [])
                    if orgs and isinstance(orgs, list) and len(orgs) > 0:
                        org = orgs[0].get("title", "") if isinstance(orgs[0], dict) else ""
                    
                    publications.append({
                        "title": r.get("title", ""),
                        "link": r.get("link", ""),
                        "organisation": org,
                        "date": r.get("public_timestamp", "")
                    })
            
            start += batch_size
            
            if len(results) < batch_size:
                break
                
        except Exception as e:
            print(f"  Warning: Error fetching senior officials page {start}: {e}")
            break
    
    # Cache the results
    try:
        with open(UK_SENIOR_OFFICIALS_CACHE, 'w') as f:
            json.dump({"publications": publications, "timestamp": datetime.now().isoformat()}, f)
    except:
        pass
    
    print(f"  Found {len(publications)} senior officials meetings publications")
    return publications


def get_senior_officials_csv_urls_from_publication(publication_path: str) -> list:
    """
    Fetch a senior officials publication from GOV.UK Content API and extract CSV URLs.
    
    Returns list of tuples: (department, csv_url, is_senior_official)
    """
    csv_urls = []
    
    try:
        url = f"{UK_GOVUK_CONTENT_API}{publication_path}"
        response = requests.get(url, timeout=30)
        response.raise_for_status()
        data = response.json()
        
        # Get organisation name
        orgs = data.get("links", {}).get("organisations", [])
        org_name = orgs[0].get("title", "Unknown") if orgs else "Unknown"
        
        # Extract CSV URLs from the documents HTML
        details = data.get("details", {})
        documents = details.get("documents", [])
        
        for doc in documents:
            urls = re.findall(r'href="(https://assets\.publishing\.service\.gov\.uk/[^"]+\.csv)"', doc)
            for csv_url in urls:
                csv_lower = csv_url.lower()
                # Only include meetings CSVs, not expenses/hospitality
                if 'meeting' in csv_lower and 'expense' not in csv_lower and 'hospitality' not in csv_lower:
                    csv_urls.append((org_name, csv_url, True))  # True = senior official
        
    except Exception:
        pass
    
    return csv_urls


# ============================================================================
# UK MEETINGS - FAST INDEX-BASED SEARCH
# ============================================================================

_uk_index_cache = {"data": None, "loaded": False}
UK_INDEX_URL = "https://raw.githubusercontent.com/your-username/telemachus/main/uk_meetings_index.json"


def load_uk_index():
    """Load the pre-built UK meetings index."""
    
    if _uk_index_cache["loaded"]:
        return _uk_index_cache["data"]
    
    # Try local file first
    local_path = Path(__file__).parent / "uk_meetings_index.json"
    
    if local_path.exists():
        try:
            with open(local_path, 'r', encoding='utf-8') as f:
                _uk_index_cache["data"] = json.load(f)
                _uk_index_cache["loaded"] = True
                print(f"  Loaded UK index from local file ({len(_uk_index_cache['data']['meetings'])} meetings)")
                return _uk_index_cache["data"]
        except Exception as e:
            print(f"  Error loading local index: {e}")
    
    # Try remote URL
    try:
        response = requests.get(UK_INDEX_URL, timeout=30)
        response.raise_for_status()
        _uk_index_cache["data"] = response.json()
        _uk_index_cache["loaded"] = True
        print(f"  Loaded UK index from remote ({len(_uk_index_cache['data']['meetings'])} meetings)")
        return _uk_index_cache["data"]
    except Exception as e:
        print(f"  Could not load UK index: {e}")
        return None


def search_uk_index(search_term: str) -> dict:
    """
    Fast search of pre-built UK meetings index.
    
    This is instant because it searches a local JSON file instead of
    making hundreds of API calls.
    """
    print(f"Searching UK meetings index for '{search_term}'...")
    
    index = load_uk_index()
    
    if not index:
        print("  No index available, falling back to live search")
        return None
    
    search_lower = search_term.lower()
    meetings = index["meetings"]
    
    # Find matching meetings
    matches = []
    for m in meetings:
        org = m.get("organisation", "").lower()
        if search_lower in org:
            matches.append(m)
    
    if not matches:
        print(f"  No meetings found for '{search_term}'")
        return None
    
    # Aggregate stats
    by_minister = {}
    by_department = {}
    by_year = {}
    
    for m in matches:
        minister = m.get("minister", "Unknown")
        dept = m.get("department", "Unknown")
        date = m.get("date", "")
        
        by_minister[minister] = by_minister.get(minister, 0) + 1
        by_department[dept] = by_department.get(dept, 0) + 1
        
        # Extract year
        year = ""
        if "/" in date:
            parts = date.split("/")
            year = parts[2] if len(parts) > 2 and len(parts[2]) == 4 else parts[0] if len(parts[0]) == 4 else ""
        elif "-" in date:
            year = date[:4]
        if year:
            by_year[year] = by_year.get(year, 0) + 1
    
    result = {
        "search_term": search_term,
        "meetings": matches,
        "meetings_count": len(matches),
        "meeting_count": len(matches),
        "by_minister": dict(sorted(by_minister.items(), key=lambda x: -x[1])),
        "by_department": dict(sorted(by_department.items(), key=lambda x: -x[1])),
        "by_year": dict(sorted(by_year.items(), key=lambda x: x[0], reverse=True)),
        "data_coverage": index["metadata"].get("coverage", "2012-present"),
        "index_date": index["metadata"].get("created", "Unknown"),
        "note": "UK ministerial and senior officials meetings from pre-built index. Fast search."
    }
    
    print(f"  Found {len(matches)} meetings")
    
    return result


def search_uk_ministerial_meetings(search_term: str, use_index: bool = True) -> dict:
    """
    Search UK ministerial meetings for an organisation.
    
    By default, uses a pre-built index for instant results (~25ms).
    The index contains ~13,000 meetings from the last 2 years.
    
    Args:
        search_term: Company or organisation name to search for
        use_index: If True (default), use fast pre-built index. If False, search live.
    """
    
    # Use index for fast search
    if use_index:
        result = search_uk_index(search_term)
        if result:
            return result
        print("  Index not available, falling back to live search...")
    
    # Fall back to live search (slow)
    return _search_uk_ministerial_meetings_live(search_term)


def _search_uk_ministerial_meetings_live(search_term: str, months_back: int = 6) -> dict:
    """
    Live search of UK ministerial meetings (slower, ~40-80 seconds).
    
    This function dynamically discovers all ministerial meetings transparency
    publications from GOV.UK, then searches their CSV attachments.
    """
    print(f"Searching UK ministerial meetings for '{search_term}'...")
    
    all_meetings = []
    departments_searched = set()
    csvs_processed = 0
    search_lower = search_term.lower()
    
    # Calculate cutoff date
    from datetime import datetime, timedelta
    cutoff_date = (datetime.now() - timedelta(days=months_back * 30)).strftime("%Y-%m-%d")
    
    # Step 1: Discover all ministerial transparency publications
    publications = discover_uk_transparency_publications()
    
    # Filter by date (only publications from last N months)
    publications_filtered = [
        p for p in publications 
        if p.get("date", "0000-00-00") >= cutoff_date
    ]
    
    # Sort by date (most recent first)
    publications_sorted = sorted(
        publications_filtered, 
        key=lambda x: x.get("date", ""), 
        reverse=True
    )
    
    print(f"  Processing {len(publications_sorted)} publications from last {months_back} months...")
    
    # Step 2: For each publication, get CSV URLs and search them
    # We'll cache the CSV URL discovery per publication
    csv_urls_cache_file = UK_MEETINGS_CACHE_DIR / "csv_urls_cache.json"
    csv_urls_cache = {}
    
    if csv_urls_cache_file.exists():
        try:
            cache_age = datetime.now().timestamp() - csv_urls_cache_file.stat().st_mtime
            if cache_age < 86400:  # 24 hours
                with open(csv_urls_cache_file, 'r') as f:
                    csv_urls_cache = json.load(f)
        except:
            pass
    
    all_csv_urls = []
    
    for pub in publications_sorted:
        pub_path = pub.get("link", "")
        if not pub_path:
            continue
            
        # Check cache first
        if pub_path in csv_urls_cache:
            csv_urls = csv_urls_cache[pub_path]
        else:
            csv_urls = get_csv_urls_from_publication(pub_path)
            csv_urls_cache[pub_path] = csv_urls
        
        all_csv_urls.extend(csv_urls)
    
    # Save CSV URLs cache
    try:
        with open(csv_urls_cache_file, 'w') as f:
            json.dump(csv_urls_cache, f)
    except:
        pass
    
    # Deduplicate CSV URLs (same URL might appear in multiple publications)
    seen_urls = set()
    unique_csv_urls = []
    for dept, url in all_csv_urls:
        if url not in seen_urls:
            seen_urls.add(url)
            unique_csv_urls.append((dept, url))
    
    print(f"  Found {len(unique_csv_urls)} unique CSV files to search")
    
    # Step 3: Download and search each CSV
    for dept, csv_url in unique_csv_urls:
        try:
            cache_key = re.sub(r'[^\w]', '_', csv_url.split('/')[-1])[:100]
            csv_path = download_uk_meetings_csv(csv_url, cache_key)
            matches = parse_uk_csv_for_matches(csv_path, search_lower, dept)
            all_meetings.extend(matches)
            departments_searched.add(dept)
            csvs_processed += 1
        except Exception as e:
            continue  # Skip unavailable files
    
    print(f"  Processed {csvs_processed} ministerial CSV files")
    
    if not all_meetings:
        print(f"  No meetings found matching '{search_term}'")
        return None
    
    # Deduplicate meetings (same minister, date, org)
    seen = set()
    unique_meetings = []
    for m in all_meetings:
        key = (m.get("minister", ""), m.get("date", ""), m.get("organisation", ""))
        if key not in seen:
            seen.add(key)
            unique_meetings.append(m)
    
    all_meetings = unique_meetings
    
    # Sort by date (most recent first) - handle various date formats
    def parse_date_for_sort(m):
        date = m.get("date", "")
        # Try DD/MM/YYYY format
        if "/" in date:
            parts = date.split("/")
            if len(parts) == 3:
                try:
                    return f"{parts[2]}-{parts[1].zfill(2)}-{parts[0].zfill(2)}"
                except:
                    pass
        # Try YYYY-MM-DD format
        if "-" in date and len(date) >= 10:
            return date[:10]
        return "0000-00-00"
    
    all_meetings.sort(key=parse_date_for_sort, reverse=True)
    
    # Aggregate statistics
    by_minister = {}
    by_department = {}
    by_year = {}
    
    for m in all_meetings:
        minister = m.get("minister", "Unknown")
        dept = m.get("department", "Unknown")
        date = m.get("date", "")
        
        # Parse year from date (various formats: DD/MM/YYYY, YYYY-MM-DD, etc.)
        year = ""
        if "/" in date:
            parts = date.split("/")
            if len(parts) == 3:
                year = parts[2] if len(parts[2]) == 4 else parts[0]
        elif "-" in date:
            year = date[:4]
        
        by_minister[minister] = by_minister.get(minister, 0) + 1
        by_department[dept] = by_department.get(dept, 0) + 1
        if year:
            by_year[year] = by_year.get(year, 0) + 1
    
    # Calculate date range from the data
    years_list = sorted(by_year.keys()) if by_year else []
    date_range = f"{min(years_list)}-{max(years_list)}" if years_list else "N/A"
    
    result = {
        "search_term": search_term,
        "meetings": all_meetings,
        "meetings_count": len(all_meetings),
        "meeting_count": len(all_meetings),  # Alias for compatibility
        "departments_searched": list(departments_searched),
        "csvs_processed": csvs_processed,
        "by_minister": dict(sorted(by_minister.items(), key=lambda x: -x[1])),
        "by_department": dict(sorted(by_department.items(), key=lambda x: -x[1])),
        "by_year": dict(sorted(by_year.items(), key=lambda x: x[0], reverse=True)),
        "data_coverage": f"Last {months_back} months",
        "date_range": date_range,
        "note": "UK ministerial and senior officials meetings from GOV.UK transparency publications (dynamically discovered). Does not include lobbying expenditure."
    }
    
    if include_senior_officials:
        result["senior_officials_meetings"] = senior_officials_count
    
    print(f"  Found {result['meetings_count']} total meetings")
    if include_senior_officials and senior_officials_count > 0:
        print(f"    (including {senior_officials_count} senior officials meetings)")
    print(f"  Departments: {len(departments_searched)}")
    if by_minister:
        print(f"  Top ministers: {', '.join(list(by_minister.keys())[:3])}")
    
    return result


def search_uk_senior_officials_meetings(search_term: str, months_back: int = 6) -> dict:
    """
    Search UK senior officials meetings for an organisation.
    
    This function dynamically discovers senior officials meetings transparency
    publications from GOV.UK, then searches their CSV attachments.
    
    Senior officials include: Permanent Secretaries, Directors General,
    Finance Directors, Commercial Directors, and other SCS2+ grade officials.
    
    Data source: GOV.UK senior officials transparency publications
    Published quarterly by each government department.
    
    Args:
        search_term: Company or organisation name to search for
        months_back: How many months of data to search (default 6)
    
    The discovery is cached for 24 hours to avoid repeated API calls.
    """
    print(f"Searching UK senior officials meetings for '{search_term}'...")
    
    all_meetings = []
    departments_searched = set()
    csvs_processed = 0
    search_lower = search_term.lower()
    
    # Calculate cutoff date
    from datetime import datetime, timedelta
    cutoff_date = (datetime.now() - timedelta(days=months_back * 30)).strftime("%Y-%m-%d")
    
    # Step 1: Discover senior officials publications
    publications = discover_uk_senior_officials_publications()
    
    # Filter by date
    publications_filtered = [
        p for p in publications
        if p.get("date", "0000-00-00") >= cutoff_date
    ]
    
    publications_sorted = sorted(
        publications_filtered,
        key=lambda x: x.get("date", ""),
        reverse=True
    )
    
    print(f"  Processing {len(publications_sorted)} publications from last {months_back} months...")
    
    # Step 2: Get CSV URLs with caching
    csv_urls_cache_file = UK_MEETINGS_CACHE_DIR / "senior_officials_csv_urls_cache.json"
    csv_urls_cache = {}
    
    if csv_urls_cache_file.exists():
        try:
            cache_age = datetime.now().timestamp() - csv_urls_cache_file.stat().st_mtime
            if cache_age < 86400:
                with open(csv_urls_cache_file, 'r') as f:
                    csv_urls_cache = json.load(f)
        except:
            pass
    
    all_csv_urls = []
    for pub in publications_sorted:
        pub_path = pub.get("link", "")
        if not pub_path:
            continue
        
        if pub_path in csv_urls_cache:
            csv_urls = csv_urls_cache[pub_path]
        else:
            csv_urls = get_senior_officials_csv_urls_from_publication(pub_path)
            csv_urls_cache[pub_path] = csv_urls
        
        all_csv_urls.extend(csv_urls)
    
    # Save cache
    try:
        with open(csv_urls_cache_file, 'w') as f:
            json.dump(csv_urls_cache, f)
    except:
        pass
    
    # Deduplicate
    seen_urls = set()
    unique_csv_urls = []
    for item in all_csv_urls:
        if len(item) == 3:
            dept, url, _ = item
        else:
            dept, url = item
        if url not in seen_urls:
            seen_urls.add(url)
            unique_csv_urls.append((dept, url))
    
    print(f"  Found {len(unique_csv_urls)} unique CSV files to search")
    
    # Step 3: Download and search each CSV
    for dept, csv_url in unique_csv_urls:
        try:
            cache_key = re.sub(r'[^\w]', '_', csv_url.split('/')[-1])[:100]
            csv_path = download_uk_meetings_csv(csv_url, cache_key)
            matches = parse_uk_csv_for_matches(csv_path, search_lower, dept)
            all_meetings.extend(matches)
            departments_searched.add(dept)
            csvs_processed += 1
        except Exception:
            continue
    
    print(f"  Processed {csvs_processed} CSV files")
    
    if not all_meetings:
        print(f"  No meetings found matching '{search_term}'")
        return None
    
    # Deduplicate meetings
    seen = set()
    unique_meetings = []
    for m in all_meetings:
        key = (m.get("minister", ""), m.get("date", ""), m.get("organisation", ""))
        if key not in seen:
            seen.add(key)
            unique_meetings.append(m)
    
    all_meetings = unique_meetings
    
    # Sort by date
    def parse_date_for_sort(m):
        date = m.get("date", "")
        if "/" in date:
            parts = date.split("/")
            if len(parts) == 3:
                try:
                    return f"{parts[2]}-{parts[1].zfill(2)}-{parts[0].zfill(2)}"
                except:
                    pass
        if "-" in date and len(date) >= 10:
            return date[:10]
        return "0000-00-00"
    
    all_meetings.sort(key=parse_date_for_sort, reverse=True)
    
    # Aggregate statistics
    by_official = {}
    by_department = {}
    by_year = {}
    
    for m in all_meetings:
        official = m.get("minister", "Unknown")  # Column is often still called "minister"
        dept = m.get("department", "Unknown")
        date = m.get("date", "")
        
        year = ""
        if "/" in date:
            parts = date.split("/")
            if len(parts) == 3:
                year = parts[2] if len(parts[2]) == 4 else parts[0]
        elif "-" in date:
            year = date[:4]
        
        by_official[official] = by_official.get(official, 0) + 1
        by_department[dept] = by_department.get(dept, 0) + 1
        if year:
            by_year[year] = by_year.get(year, 0) + 1
    
    result = {
        "search_term": search_term,
        "meetings": all_meetings,
        "meetings_count": len(all_meetings),
        "meeting_count": len(all_meetings),
        "departments_searched": list(departments_searched),
        "csvs_processed": csvs_processed,
        "by_official": dict(sorted(by_official.items(), key=lambda x: -x[1])),
        "by_minister": dict(sorted(by_official.items(), key=lambda x: -x[1])),  # Alias
        "by_department": dict(sorted(by_department.items(), key=lambda x: -x[1])),
        "by_year": dict(sorted(by_year.items(), key=lambda x: x[0], reverse=True)),
        "data_coverage": f"Last {months_back} months",
        "note": "UK senior officials (Permanent Secretaries, DGs, SCS2+) meetings from GOV.UK transparency publications. Does not include lobbying expenditure."
    }
    
    print(f"  Found {result['meetings_count']} senior officials meetings")
    print(f"  Departments: {len(departments_searched)}")
    if by_official:
        print(f"  Top officials: {', '.join(list(by_official.keys())[:3])}")
    
    return result


# ============================================================================
# AUSTRIA - Lobbying- und Interessenvertretungs-Register
# ============================================================================

AUSTRIA_REGISTER_URL = "https://lobbyreg.justiz.gv.at/edikte/ir/iredi18.nsf/liste!OpenForm&subf=a"
AUSTRIA_DETAIL_URL = "https://lobbyreg.justiz.gv.at/edikte/ir/iredi18.nsf/alldoc/{doc_id}!OpenDocument"
AUSTRIA_CACHE_DIR = Path.home() / ".cache" / "eu_lobbying" / "austria"


def search_austria_register(search_term: str) -> dict:
    """
    Search Austrian Lobbying- und Interessenvertretungs-Register.
    
    The Austrian register uses a Lotus Notes/Domino system. We scrape the
    public alphabetical list to find matching organisations.
    
    Register categories:
    - A1: Lobbying companies (lobby for clients)
    - A2: Client information (not public)
    - B: Companies with in-house lobbyists
    - C: Self-governing bodies
    - D: Interest groups/associations
    
    Data source: https://lobbyreg.justiz.gv.at
    Administered by: Austrian Federal Ministry of Justice
    """
    print(f"Searching Austrian register for '{search_term}'...")
    
    AUSTRIA_CACHE_DIR.mkdir(parents=True, exist_ok=True)
    cache_file = AUSTRIA_CACHE_DIR / "register_list.html"
    
    # Cache the list page for 24 hours
    if cache_file.exists() and (datetime.now().timestamp() - cache_file.stat().st_mtime) < 86400:
        with open(cache_file, 'r', encoding='utf-8', errors='replace') as f:
            html_content = f.read()
    else:
        print("  Fetching Austrian register list...")
        try:
            response = requests.get(
                AUSTRIA_REGISTER_URL,
                headers={"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36"},
                timeout=60
            )
            response.raise_for_status()
            html_content = response.text
            with open(cache_file, 'w', encoding='utf-8') as f:
                f.write(html_content)
        except Exception as e:
            print(f"  Error fetching Austrian register: {e}")
            return None
    
    # Parse the table data using regex (avoiding heavy HTML parser dependency)
    # Pattern matches table rows with: name, register number, category, lobbyists, date
    td_pattern = re.compile(r'<td[^>]*>(.*?)</td>', re.DOTALL)
    cells = td_pattern.findall(html_content)
    
    # Group cells into rows (6 cells per row: count, name, regnum, category, lobbyists, date)
    matches = []
    search_lower = search_term.lower()
    
    i = 0
    while i < len(cells) - 5:
        # Skip the count cell (contains JavaScript)
        if 'count()' in cells[i]:
            i += 1
            continue
        
        name_cell = cells[i] if i < len(cells) else ""
        regnum_cell = cells[i+1] if i+1 < len(cells) else ""
        category_cell = cells[i+2] if i+2 < len(cells) else ""
        lobbyists_cell = cells[i+3] if i+3 < len(cells) else ""
        date_cell = cells[i+4] if i+4 < len(cells) else ""
        
        # Extract clean text
        name = re.sub(r'<[^>]+>', '', name_cell).strip()
        
        # Check if this row matches our search
        if search_lower in name.lower():
            # Extract register number from link
            regnum_match = re.search(r'>(LIVR-\d+)</a>', regnum_cell)
            regnum = regnum_match.group(1) if regnum_match else ""
            
            # Extract document ID for detail link
            doc_id_match = re.search(r'alldoc/([a-f0-9]+)!', regnum_cell)
            doc_id = doc_id_match.group(1) if doc_id_match else ""
            
            # Clean category
            category = re.sub(r'<[^>]+>', '', category_cell).strip()
            
            # Clean lobbyists (replace <br> with comma)
            lobbyists = re.sub(r'<br\s*/?>', ', ', lobbyists_cell)
            lobbyists = re.sub(r'<[^>]+>', '', lobbyists).strip()
            lobbyists = re.sub(r',\s*$', '', lobbyists)  # Remove trailing comma
            
            # Clean date
            date = re.sub(r'<[^>]+>', '', date_cell).strip()
            
            matches.append({
                "name": name,
                "register_number": regnum,
                "doc_id": doc_id,
                "category": category,
                "category_description": {
                    "A1": "Lobbying company",
                    "A2": "Client info (not public)",
                    "B": "Company with in-house lobbyists",
                    "C": "Self-governing body",
                    "D": "Interest group/association"
                }.get(category, category),
                "lobbyists": lobbyists,
                "last_update": date,
                "detail_url": AUSTRIA_DETAIL_URL.format(doc_id=doc_id) if doc_id else ""
            })
        
        i += 5  # Move to next row
    
    if not matches:
        print(f"  No matches found for '{search_term}' in Austrian register")
        return None
    
    print(f"  Found {len(matches)} matching entries")
    
    # Return aggregated results
    result = {
        "search_term": search_term,
        "entries": matches,
        "entry_count": len(matches),
        "by_category": {},
        "data_coverage": "2013-present",
        "note": "Austrian register (lobbyreg.justiz.gv.at). Financial data only shows if costs >€100,000."
    }
    
    # Count by category
    for entry in matches:
        cat = entry.get("category", "Unknown")
        result["by_category"][cat] = result["by_category"].get(cat, 0) + 1
    
    return result


def get_austria_detail(doc_id: str) -> dict:
    """Fetch detailed information for an Austrian register entry."""
    if not doc_id:
        return None
    
    url = AUSTRIA_DETAIL_URL.format(doc_id=doc_id)
    
    try:
        response = requests.get(
            url,
            headers={"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36"},
            timeout=30
        )
        response.raise_for_status()
        html = response.text
        
        # Extract key fields using regex
        detail = {"url": url}
        
        # Activity area (Tätigkeitsbereich)
        activity_match = re.search(r'Tätigkeitsbereich.*?<td[^>]*>(.*?)</td>', html, re.DOTALL)
        if activity_match:
            detail["activity_area"] = re.sub(r'<[^>]+>', ' ', activity_match.group(1)).strip()
        
        # Website
        website_match = re.search(r'Website.*?<a[^>]*href="([^"]+)"', html, re.DOTALL)
        if website_match:
            detail["website"] = website_match.group(1)
        
        # Lobbying costs > 100,000€
        cost_match = re.search(r'Lobbying-Aufwand.*?>([^<]*100\.000)', html, re.DOTALL | re.IGNORECASE)
        if cost_match:
            detail["high_cost_flag"] = True
        
        # Company register number
        firmenbuch_match = re.search(r'Firmenbuchnummer.*?<td[^>]*>([^<]+)</td>', html, re.DOTALL)
        if firmenbuch_match:
            detail["company_register_number"] = firmenbuch_match.group(1).strip()
        
        return detail
        
    except Exception as e:
        return None


# ============================================================================
# CATALONIA - Registre de grups d'interès de Catalunya
# ============================================================================

CATALONIA_API_URL = "https://analisi.transparenciacatalunya.cat/resource/gwpn-de62.json"


def search_catalonia_register(search_term: str) -> dict:
    """
    Search Catalonia's Interest Group Register (Registre de grups d'interès de Catalunya).
    
    The Catalan register covers lobbying activities before the Generalitat de Catalunya
    (regional government), Parliament, local administrations, and other public bodies.
    
    Categories:
    - I: Consultancy and advisory services
    - II: Business sector and associations  
    - III: Non-governmental organisations
    - IV: Think tanks, research centres, academic institutions
    - V: Citizens' organisations
    - VI: Other entities
    
    Data source: https://analisi.transparenciacatalunya.cat
    Administered by: Generalitat de Catalunya
    """
    print(f"Searching Catalonia register for '{search_term}'...")
    
    # Socrata API with case-insensitive search
    params = {
        "$where": f"lower(nom) like '%{search_term.lower()}%'",
        "$limit": 100
    }
    
    try:
        response = requests.get(CATALONIA_API_URL, params=params, timeout=30)
        response.raise_for_status()
        data = response.json()
    except Exception as e:
        print(f"  Error searching Catalonia register: {e}")
        return None
    
    if not data:
        print(f"  No matches found for '{search_term}' in Catalonia register")
        return None
    
    print(f"  Found {len(data)} matching entries")
    
    entries = []
    for item in data:
        # Parse annual business volume
        volume = item.get("volum_anual_negocis", "0")
        try:
            volume_int = int(float(volume)) if volume else 0
        except:
            volume_int = 0
        
        # Parse budget
        budget = item.get("pressupost", "0")
        try:
            budget_int = int(float(budget)) if budget else 0
        except:
            budget_int = 0
        
        entries.append({
            "name": item.get("nom", ""),
            "id": item.get("identificador", ""),
            "registration_date": item.get("data_alta", "")[:10] if item.get("data_alta") else "",
            "entity_type": item.get("tipus_grup", ""),
            "category": item.get("categoria_registre", ""),
            "subcategory": item.get("subcategoria_registre", ""),
            "purpose": item.get("finalitat", ""),
            "areas_of_interest": item.get("ambits_interes", ""),
            "annual_volume": volume_int,
            "annual_volume_formatted": f"€{volume_int:,}" if volume_int > 0 else "Not disclosed",
            "budget": budget_int,
            "public_funds": item.get("fons_p_blics", "0"),
            "email": item.get("email", ""),
            "website": item.get("pagina_web", ""),
            "province": item.get("provincia", ""),
            "municipality": item.get("muni_esp", ""),
            "scope": item.get("ambits_registre", ""),
            "activities": item.get("propostes_normatives", ""),
        })
    
    # Aggregate by category
    by_category = {}
    total_volume = 0
    for entry in entries:
        cat = entry.get("category", "Unknown")
        # Simplify category name
        cat_short = cat.replace("Categoria ", "").split(".")[0] if "Categoria" in cat else cat
        by_category[cat_short] = by_category.get(cat_short, 0) + 1
        total_volume += entry.get("annual_volume", 0)
    
    result = {
        "search_term": search_term,
        "entries": entries,
        "entry_count": len(entries),
        "by_category": by_category,
        "total_annual_volume": total_volume,
        "total_volume_formatted": f"€{total_volume:,}" if total_volume > 0 else "Not disclosed",
        "data_coverage": "2016-present",
        "note": "Catalonia Interest Group Register (analisi.transparenciacatalunya.cat). Covers lobbying before Generalitat, Parliament, and local administrations."
    }
    
    return result


# ============================================================================
# FINLAND - Avoimuusrekisteri (Transparency Register)
# ============================================================================

FINLAND_API_URL = "https://public.api.avoimuusrekisteri.fi"

# Cache for Finland registrations (refreshed every 24 hours)
_finland_cache = {"data": None, "timestamp": 0}
FINLAND_CACHE_DURATION = 86400  # 24 hours


def get_finland_registrations() -> list:
    """
    Fetch all registrations from Finland's Transparency Register.
    Results are cached for 24 hours since the full list must be fetched.
    """
    import time
    
    now = time.time()
    if _finland_cache["data"] and (now - _finland_cache["timestamp"]) < FINLAND_CACHE_DURATION:
        return _finland_cache["data"]
    
    try:
        response = requests.get(f"{FINLAND_API_URL}/open-data-register-notification", timeout=30)
        response.raise_for_status()
        data = response.json()
        _finland_cache["data"] = data
        _finland_cache["timestamp"] = now
        return data
    except Exception as e:
        print(f"  Error fetching Finland registrations: {e}")
        return []


def get_finland_activities(company_id: str) -> list:
    """
    Fetch activity disclosures for a specific company from Finland's Transparency Register.
    """
    try:
        response = requests.get(f"{FINLAND_API_URL}/open-data-activity-notification/company/{company_id}", timeout=30)
        response.raise_for_status()
        return response.json()
    except Exception as e:
        return []


def search_finland_register(search_term: str) -> dict:
    """
    Search Finland's Transparency Register (Avoimuusrekisteri).
    
    The Finnish register covers lobbying activities targeting:
    - Parliament (Eduskunta)
    - Ministries
    
    Data includes:
    - Registration details
    - Activity disclosures (submitted twice yearly)
    - Lobbying topics and targets contacted
    - From 2026: Financial information
    
    Data source: https://avoimuusrekisteri.fi
    Administered by: National Audit Office of Finland (VTV)
    """
    print(f"Searching Finland register for '{search_term}'...")
    
    # Get all registrations and filter locally
    all_registrations = get_finland_registrations()
    
    if not all_registrations:
        print(f"  Could not fetch Finland registrations")
        return None
    
    search_lower = search_term.lower()
    matches = []
    
    for reg in all_registrations:
        company_name = reg.get("companyName", "")
        # Check main name
        if search_lower in company_name.lower():
            matches.append(reg)
            continue
        # Check supplementary names
        for supp in reg.get("supplementaryCompanyNames", []):
            if search_lower in supp.get("title", "").lower():
                matches.append(reg)
                break
    
    if not matches:
        print(f"  No matches found for '{search_term}' in Finland register")
        return None
    
    print(f"  Found {len(matches)} matching registration(s)")
    
    entries = []
    total_activities = 0
    
    for reg in matches:
        company_id = reg.get("companyId", "")
        
        # Fetch activity disclosures for this company
        activities = get_finland_activities(company_id) if company_id else []
        total_activities += len(activities)
        
        # Extract topics from activities
        topics = set()
        targets_contacted = set()
        for act in activities:
            for topic in act.get("topics", []):
                topic_text = topic.get("contactTopicOther") or topic.get("contactTopicProject") or ""
                if topic_text and isinstance(topic_text, str):
                    topics.add(topic_text[:100] if len(topic_text) > 100 else topic_text)
        
        # Get memberships
        memberships = [m.get("title", "") for m in reg.get("memberships", [])]
        
        entries.append({
            "name": reg.get("companyName", ""),
            "company_id": company_id,
            "diary_number": reg.get("diaryNumber", ""),
            "registration_date": reg.get("registrationDate", ""),
            "main_industry": reg.get("mainIndustry", ""),
            "description": reg.get("description", ""),
            "memberships": memberships,
            "supplementary_names": [s.get("title", "") for s in reg.get("supplementaryCompanyNames", [])],
            "activity_count": len(activities),
            "topics": list(topics)[:10],  # Limit to 10 topics
            "ethical_commitment": reg.get("commitToEthicalLobbying"),
            "exited": reg.get("exited", False),
        })
    
    result = {
        "search_term": search_term,
        "entries": entries,
        "entry_count": len(entries),
        "total_activities": total_activities,
        "data_coverage": "2024-present (financial data from July 2026)",
        "note": "Finland Transparency Register (avoimuusrekisteri.fi). Financial data available from 2026."
    }
    
    return result


# =============================================================================
# SLOVENIA (KPK - Commission for the Prevention of Corruption)
# =============================================================================

SLOVENIA_REGISTER_URL = "https://www.kpk-rs.si/sl/lobiranje-22/register-lobistov"


def get_slovenia_lobbyists() -> list:
    """
    Fetch all registered lobbyists from Slovenia's KPK register.
    
    The Slovenian register tracks individual lobbyists (natural persons), not companies.
    Each lobbyist entry includes:
    - Name
    - Employer/company (if applicable)
    - Fields of interest they lobby on
    - Contact information
    
    Data source: https://www.kpk-rs.si/sl/lobiranje-22/register-lobistov
    Administered by: Commission for the Prevention of Corruption (KPK)
    """
    from bs4 import BeautifulSoup
    
    cache_file = get_cache_dir() / "slovenia_lobbyists.html"
    cache_max_age = 24 * 60 * 60  # 24 hours
    
    # Use cache if fresh
    if cache_file.exists():
        age = (datetime.now().timestamp() - cache_file.stat().st_mtime)
        if age < cache_max_age:
            html_content = cache_file.read_text(encoding='utf-8')
        else:
            html_content = None
    else:
        html_content = None
    
    # Fetch if no cache
    if html_content is None:
        try:
            headers = {
                'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36',
                'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8',
                'Accept-Language': 'en-US,en;q=0.5',
            }
            response = requests.get(SLOVENIA_REGISTER_URL, headers=headers, timeout=30)
            response.raise_for_status()
            html_content = response.text
            cache_file.write_text(html_content, encoding='utf-8')
        except Exception as e:
            print(f"  Error fetching Slovenia register: {e}")
            return []
    
    # Parse HTML
    soup = BeautifulSoup(html_content, 'html.parser')
    lobbyists = []
    
    # Find all lobbyist entries - they're in divs with class 'flex flex-column gap-3'
    # Structure: <strong>Name</strong> then <p> with fields, then <ul> with contact info
    
    # Find all strong tags that look like names (contain comma)
    for strong in soup.find_all('strong'):
        name = strong.get_text(strip=True)
        
        # Skip non-name entries
        if ',' not in name:
            continue
        if any(skip in name.lower() for skip in ['lobist', 'register', 'sankcij', 'komisija']):
            continue
        
        entry = {
            'name': name,
            'company': '',
            'fields_of_interest': [],
            'address': '',
            'city': '',
            'email': '',
        }
        
        # Get parent div
        parent = strong.find_parent('div', class_=lambda x: x and 'flex' in x)
        if not parent:
            parent = strong.parent
        
        if parent:
            # Find fields of interest - in <p> tag with middots
            fields_p = parent.find('p', class_='m-0')
            if fields_p:
                fields_text = fields_p.get_text()
                # Split by middot (·) and clean
                fields = [f.strip() for f in fields_text.split('·') if f.strip()]
                entry['fields_of_interest'] = fields
            
            # Find contact info - in <ul> with <li> items
            contact_ul = parent.find('ul')
            if contact_ul:
                for li in contact_ul.find_all('li'):
                    li_text = li.get_text(strip=True)
                    
                    # Check for email
                    if '@' in li_text:
                        entry['email'] = li_text
                    # Check for company (contains D.O.O., S.P., etc.)
                    elif any(s in li_text.upper() for s in ['D.O.O.', 'D.O.O', 'S.P.', 'S.P', 'D. O. O.']):
                        entry['company'] = li_text
                    # Check for postal code (4 digits)
                    elif re.search(r'^\d{4}\s', li_text):
                        entry['city'] = li_text
                    # Otherwise likely address
                    elif re.search(r'\d', li_text) and not entry['address']:
                        entry['address'] = li_text
        
        if entry['name']:
            lobbyists.append(entry)
    
    # Remove duplicates by name
    seen_names = set()
    unique_lobbyists = []
    for l in lobbyists:
        if l['name'] not in seen_names:
            seen_names.add(l['name'])
            unique_lobbyists.append(l)
    
    return unique_lobbyists


def search_slovenia_register(search_term: str) -> dict:
    """
    Search Slovenia's Lobbying Register (Register lobistov).
    
    IMPORTANT: Slovenia's register tracks INDIVIDUAL LOBBYISTS, not companies.
    Search matches against:
    - Lobbyist name
    - Company/employer name
    - Fields of interest
    
    To find corporate lobbying, search for:
    - Known lobbyist names
    - PR/PA firms (e.g., "Propiar", "Herman", "Bonorum")
    - Industry terms in fields of interest
    
    Data source: https://www.kpk-rs.si/sl/lobiranje-22/register-lobistov
    Administered by: Commission for the Prevention of Corruption (KPK)
    
    Note: This register shows WHO can lobby, not WHAT companies are lobbying.
    To track corporate influence, you need to cross-reference:
    - Lobbyist employer information
    - Interest organization reports (submitted separately to KPK)
    """
    print(f"Searching Slovenia register for '{search_term}'...")
    
    lobbyists = get_slovenia_lobbyists()
    
    if not lobbyists:
        print(f"  Could not fetch Slovenia lobbyist data")
        return None
    
    search_lower = search_term.lower()
    matches = []
    
    for lobbyist in lobbyists:
        # Search in name
        if search_lower in lobbyist.get('name', '').lower():
            matches.append(lobbyist)
            continue
            
        # Search in company
        if search_lower in lobbyist.get('company', '').lower():
            matches.append(lobbyist)
            continue
            
        # Search in fields of interest
        for field in lobbyist.get('fields_of_interest', []):
            if search_lower in field.lower():
                matches.append(lobbyist)
                break
    
    if not matches:
        print(f"  No matches found for '{search_term}' in Slovenia register")
        return None
    
    print(f"  Found {len(matches)} matching lobbyist(s)")
    
    # Organize results
    entries = []
    fields_summary = {}
    
    for m in matches:
        entries.append({
            'name': m.get('name', ''),
            'company': m.get('company', ''),
            'fields_of_interest': m.get('fields_of_interest', []),
            'address': m.get('address', ''),
            'city': m.get('city', ''),
            'email': m.get('email', ''),
        })
        
        # Count field occurrences
        for field in m.get('fields_of_interest', []):
            fields_summary[field] = fields_summary.get(field, 0) + 1
    
    # Sort fields by frequency
    top_fields = sorted(fields_summary.items(), key=lambda x: -x[1])[:10]
    
    result = {
        "search_term": search_term,
        "entries": entries,
        "entry_count": len(entries),
        "top_fields": top_fields,
        "total_registered": len(lobbyists),
        "data_coverage": "2010-present",
        "note": "Slovenia Register of Lobbyists (KPK). Lists individual lobbyists, not companies. Cross-reference with interest organization reports for company activities."
    }
    
    return result


def create_excel_report(eu_data: dict, fr_data: dict, de_data: dict, ie_data: dict, uk_data: dict, at_data: dict, cat_data: dict, fi_data: dict, si_data: dict = None, uk_officials_data: dict = None, output_path: str = None, org_name: str = None):
    """Create combined Excel report."""
    wb = Workbook()
    
    # Styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill_eu = PatternFill("solid", fgColor="2F5496")  # Blue for EU
    header_fill_fr = PatternFill("solid", fgColor="C00000")  # Red for France
    header_fill_de = PatternFill("solid", fgColor="FFD700")  # Gold for Germany
    header_fill_ie = PatternFill("solid", fgColor="169B62")  # Green for Ireland
    header_fill_uk = PatternFill("solid", fgColor="012169")  # UK Blue
    header_fill_uk_officials = PatternFill("solid", fgColor="4A4A8A")  # UK Officials Purple-ish
    header_fill_at = PatternFill("solid", fgColor="ED2939")  # Austria Red
    header_fill_cat = PatternFill("solid", fgColor="FCDD09")  # Catalonia Yellow (Senyera)
    header_fill_fi = PatternFill("solid", fgColor="003580")  # Finland Blue
    header_fill_si = PatternFill("solid", fgColor="005DA4")  # Slovenia Blue
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell_align = Alignment(vertical="top", wrap_text=True)
    border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    
    # === SUMMARY SHEET ===
    ws = wb.active
    ws.title = "Summary"
    
    ws['A1'] = f"European Lobbying Report: {org_name}"
    ws['A1'].font = Font(bold=True, size=16)
    ws.merge_cells('A1:D1')
    
    ws['A2'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    ws['A2'].font = Font(italic=True)
    ws.merge_cells('A2:D2')
    
    row = 4
    
    # EU Summary
    if eu_data:
        ws.cell(row=row, column=1, value="EU TRANSPARENCY REGISTER").font = Font(bold=True, size=14, color="2F5496")
        row += 1
        
        regs = eu_data.get("registrations", [])
        meetings = eu_data.get("meetings", [])
        latest = regs[-1] if regs else {}
        
        eu_stats = [
            ("Organisation", latest.get("original_name", org_name)),
            ("EU Register ID", eu_data.get("org_id", "")),
            ("Country (HQ)", latest.get("head_country", "")),
            ("Lobbying Costs (latest)", f"€{int(latest.get('min', 0) or 0):,} - €{int(latest.get('max', 0) or 0):,}" if latest.get('min') or latest.get('max') else "Not disclosed"),
            ("Staff", f"{latest.get('members', 'N/A')} ({latest.get('members_fte', 'N/A')} FTE)"),
            ("Commission Meetings", str(len(meetings))),
            ("Data Points (snapshots)", str(len(regs))),
        ]
        
        for label, value in eu_stats:
            ws.cell(row=row, column=1, value=label).font = Font(bold=True)
            ws.cell(row=row, column=2, value=value)
            row += 1
        
        row += 1
    
    # France Summary
    if fr_data and fr_data.get("info"):
        ws.cell(row=row, column=1, value="FRANCE (HATVP)").font = Font(bold=True, size=14, color="C00000")
        row += 1
        
        info = fr_data.get("info", {})
        exercises = fr_data.get("exercises", [])
        activities = fr_data.get("activities", [])
        latest_ex = exercises[2] if len(exercises) > 2 else (exercises[0] if exercises else {})  # Skip current incomplete years
        
        fr_stats = [
            ("Organisation", info.get("denomination", org_name)),
            ("HATVP ID", fr_data.get("org_id", "")),
            ("SIREN", info.get("identifiant_national", "")),
            ("City", info.get("ville", "")),
            ("Lobbying Costs (latest)", latest_ex.get("montant_depense", "Not disclosed")),
            ("Staff (lobbyists)", latest_ex.get("nombre_salaries", "N/A")),
            ("Lobbying Activities", str(len(activities))),
            ("Years of Data", str(len([e for e in exercises if e.get('nombre_activites', '0') != '0']))),
        ]
        
        for label, value in fr_stats:
            ws.cell(row=row, column=1, value=label).font = Font(bold=True)
            ws.cell(row=row, column=2, value=value)
            row += 1
        
        row += 1
    
    # Germany Summary
    if de_data and de_data.get("name"):
        ws.cell(row=row, column=1, value="GERMANY (Bundestag Lobbyregister)").font = Font(bold=True, size=14, color="806600")
        row += 1
        
        de_stats = [
            ("Organisation", de_data.get("name", org_name)),
            ("Register Number", de_data.get("register_number", "")),
            ("City", de_data.get("city", "")),
            ("Berlin Office", "Yes" if de_data.get("berlin_office") else "No"),
            ("Lobbying Costs", f"€{de_data.get('expenses_min', 0):,} - €{de_data.get('expenses_max', 0):,}"),
            ("Staff (FTE)", str(de_data.get("employee_fte", "N/A"))),
            ("Regulatory Projects", str(de_data.get("regulatory_projects_count", 0))),
            ("Fields of Interest", str(len(de_data.get("fields_of_interest", [])))),
        ]
        
        for label, value in de_stats:
            ws.cell(row=row, column=1, value=label).font = Font(bold=True)
            ws.cell(row=row, column=2, value=value)
            row += 1
        
        row += 1
    
    # Ireland Summary
    if ie_data and ie_data.get("name"):
        ws.cell(row=row, column=1, value="IRELAND (Lobbying.ie)").font = Font(bold=True, size=14, color="169B62")
        row += 1
        
        ie_stats = [
            ("Organisation", ie_data.get("name", org_name)),
            ("Trading Name", ie_data.get("trading_name", "") or "N/A"),
            ("Lobbying Returns", str(ie_data.get("returns_count", 0))),
            ("Lobbying Costs", "Not disclosed (not required under Irish law)"),
            ("Note", ie_data.get("note", "")),
        ]
        
        for label, value in ie_stats:
            ws.cell(row=row, column=1, value=label).font = Font(bold=True)
            ws.cell(row=row, column=2, value=value)
            row += 1
        
        row += 1
    
    # UK Summary
    if uk_data and uk_data.get("meetings"):
        ws.cell(row=row, column=1, value="UK (GOV.UK Ministerial Meetings)").font = Font(bold=True, size=14, color="012169")
        row += 1
        
        uk_stats = [
            ("Search Term", uk_data.get("search_term", org_name)),
            ("Ministerial Meetings", str(uk_data.get("meetings_count", 0))),
            ("Departments Searched", str(len(uk_data.get("departments_searched", [])))),
            ("Lobbying Costs", "Not tracked (UK publishes meetings, not expenditure)"),
            ("Top Minister", list(uk_data.get("by_minister", {}).keys())[0] if uk_data.get("by_minister") else "N/A"),
        ]
        
        for label, value in uk_stats:
            ws.cell(row=row, column=1, value=label).font = Font(bold=True)
            ws.cell(row=row, column=2, value=value)
            row += 1
        
        row += 1
    
    # UK Senior Officials Summary
    if uk_officials_data and uk_officials_data.get("meetings"):
        ws.cell(row=row, column=1, value="UK (Senior Officials Meetings)").font = Font(bold=True, size=14, color="4A4A8A")
        row += 1
        
        uk_so_stats = [
            ("Search Term", uk_officials_data.get("search_term", org_name)),
            ("Senior Officials Meetings", str(uk_officials_data.get("meetings_count", 0))),
            ("Departments", str(len(uk_officials_data.get("departments_searched", [])))),
            ("Officials Level", "Permanent Secretaries, DGs, SCS2+"),
            ("Top Official", list(uk_officials_data.get("by_official", {}).keys())[0] if uk_officials_data.get("by_official") else "N/A"),
        ]
        
        for label, value in uk_so_stats:
            ws.cell(row=row, column=1, value=label).font = Font(bold=True)
            ws.cell(row=row, column=2, value=value)
            row += 1
        
        row += 1
    
    # Austria Summary
    if at_data and at_data.get("entries"):
        ws.cell(row=row, column=1, value="AUSTRIA (Lobbying- und Interessenvertretungs-Register)").font = Font(bold=True, size=14, color="ED2939")
        row += 1
        
        entries = at_data.get("entries", [])
        categories = at_data.get("by_category", {})
        cat_desc = ", ".join([f"{k}: {v}" for k, v in categories.items()])
        
        at_stats = [
            ("Search Term", at_data.get("search_term", org_name)),
            ("Register Entries", str(at_data.get("entry_count", 0))),
            ("By Category", cat_desc if cat_desc else "N/A"),
            ("Lobbying Costs", "Only disclosed if >€100,000"),
            ("Note", "A1=Lobby firm, B=In-house, C=Self-gov, D=Interest group"),
        ]
        
        for label, value in at_stats:
            ws.cell(row=row, column=1, value=label).font = Font(bold=True)
            ws.cell(row=row, column=2, value=value)
            row += 1
        
        row += 1
    
    # Catalonia Summary
    if cat_data and cat_data.get("entries"):
        ws.cell(row=row, column=1, value="CATALONIA (Registre de grups d'interès)").font = Font(bold=True, size=14, color="DA121A")
        row += 1
        
        entries = cat_data.get("entries", [])
        categories = cat_data.get("by_category", {})
        cat_desc = ", ".join([f"{k}: {v}" for k, v in categories.items()])
        
        cat_stats = [
            ("Search Term", cat_data.get("search_term", org_name)),
            ("Register Entries", str(cat_data.get("entry_count", 0))),
            ("By Category", cat_desc if cat_desc else "N/A"),
            ("Total Annual Volume", cat_data.get("total_volume_formatted", "N/A")),
            ("Scope", "Generalitat, Parliament, local administrations"),
        ]
        
        for label, value in cat_stats:
            ws.cell(row=row, column=1, value=label).font = Font(bold=True)
            ws.cell(row=row, column=2, value=value)
            row += 1
        
        row += 1
    
    # Finland Summary
    if fi_data and fi_data.get("entries"):
        ws.cell(row=row, column=1, value="FINLAND (Avoimuusrekisteri)").font = Font(bold=True, size=14, color="003580")
        row += 1
        
        fi_stats = [
            ("Search Term", fi_data.get("search_term", org_name)),
            ("Register Entries", str(fi_data.get("entry_count", 0))),
            ("Activity Disclosures", str(fi_data.get("total_activities", 0))),
            ("Lobbying Costs", "Financial data available from July 2026"),
            ("Scope", "Parliament (Eduskunta) and Ministries"),
        ]
        
        for label, value in fi_stats:
            ws.cell(row=row, column=1, value=label).font = Font(bold=True)
            ws.cell(row=row, column=2, value=value)
            row += 1
    
    # Slovenia Summary
    if si_data and si_data.get("entries"):
        row += 1
        ws.cell(row=row, column=1, value="SLOVENIA (Register lobistov)").font = Font(bold=True, size=14, color="005DA4")
        row += 1
        
        # Get top fields summary
        top_fields_str = ", ".join([f[0] for f in si_data.get("top_fields", [])[:5]]) or "N/A"
        
        si_stats = [
            ("Search Term", si_data.get("search_term", org_name)),
            ("Lobbyists Found", str(si_data.get("entry_count", 0))),
            ("Total Registered Lobbyists", str(si_data.get("total_registered", 0))),
            ("Top Fields of Interest", top_fields_str),
            ("Note", "Lists individual lobbyists, not companies"),
        ]
        
        for label, value in si_stats:
            ws.cell(row=row, column=1, value=label).font = Font(bold=True)
            ws.cell(row=row, column=2, value=value)
            row += 1
    
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 60
    
    # === EU MEETINGS SHEET ===
    if eu_data and eu_data.get("meetings"):
        ws_meetings = wb.create_sheet("EU Commission Meetings")
        meetings = eu_data["meetings"]
        
        cols = [
            ("Date", "Date", 12),
            ("Subject", "Subject", 60),
            ("Portfolio/DG", "DG name/Portfolio", 30),
            ("Attendees", "Attending from Commission", 40),
            ("Other Lobbyists", "Other lobbyists", 40),
        ]
        
        ws_meetings['A1'] = f"EU Commission Meetings - {org_name}"
        ws_meetings['A1'].font = Font(bold=True, size=14)
        ws_meetings.merge_cells('A1:E1')
        
        for col_idx, (header, _, width) in enumerate(cols, 1):
            cell = ws_meetings.cell(row=3, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill_eu
            cell.alignment = header_align
            cell.border = border
            ws_meetings.column_dimensions[get_column_letter(col_idx)].width = width
        
        for row_idx, m in enumerate(meetings, 4):
            for col_idx, (_, field, _) in enumerate(cols, 1):
                value = m.get(field, "")[:32000] if m.get(field) else ""
                cell = ws_meetings.cell(row=row_idx, column=col_idx, value=value)
                cell.alignment = cell_align
                cell.border = border
        
        ws_meetings.freeze_panes = 'A4'
        ws_meetings.auto_filter.ref = f"A3:E{3 + len(meetings)}"
    
    # === EU COSTS HISTORY SHEET ===
    if eu_data and eu_data.get("registrations"):
        ws_costs = wb.create_sheet("EU Costs History")
        regs = eu_data["registrations"]
        
        cols = [
            ("Date", "state_date", 12),
            ("Year Start", "start_date", 12),
            ("Year End", "end_date", 12),
            ("Min €", "min", 15),
            ("Max €", "max", 15),
            ("Staff", "members", 10),
        ]
        
        ws_costs['A1'] = f"EU Lobbying Costs Over Time - {org_name}"
        ws_costs['A1'].font = Font(bold=True, size=14)
        
        for col_idx, (header, _, width) in enumerate(cols, 1):
            cell = ws_costs.cell(row=3, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill_eu
            ws_costs.column_dimensions[get_column_letter(col_idx)].width = width
        
        for row_idx, r in enumerate(regs, 4):
            for col_idx, (_, field, _) in enumerate(cols, 1):
                value = r.get(field, "")
                if field in ("min", "max") and value:
                    try:
                        value = int(float(value))
                    except:
                        pass
                ws_costs.cell(row=row_idx, column=col_idx, value=value)
    
    # === FRANCE ACTIVITIES SHEET ===
    if fr_data and fr_data.get("activities"):
        ws_fr = wb.create_sheet("France Activities")
        activities = fr_data["activities"]
        
        cols = [
            ("Date", "date_publication_activite", 12),
            ("Subject", "objet_activite", 80),
            ("ID", "identifiant_fiche", 12),
        ]
        
        ws_fr['A1'] = f"French Lobbying Activities - {org_name}"
        ws_fr['A1'].font = Font(bold=True, size=14)
        ws_fr.merge_cells('A1:C1')
        
        ws_fr['A2'] = "Source: HATVP (Haute Autorité pour la Transparence de la Vie Publique)"
        ws_fr['A2'].font = Font(italic=True, size=10)
        
        for col_idx, (header, _, width) in enumerate(cols, 1):
            cell = ws_fr.cell(row=4, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill_fr
            cell.alignment = header_align
            ws_fr.column_dimensions[get_column_letter(col_idx)].width = width
        
        for row_idx, a in enumerate(activities, 5):
            for col_idx, (_, field, _) in enumerate(cols, 1):
                value = a.get(field, "")[:32000] if a.get(field) else ""
                cell = ws_fr.cell(row=row_idx, column=col_idx, value=value)
                cell.alignment = cell_align
        
        ws_fr.freeze_panes = 'A5'
        ws_fr.auto_filter.ref = f"A4:C{4 + len(activities)}"
    
    # === FRANCE COSTS HISTORY SHEET ===
    if fr_data and fr_data.get("exercises"):
        ws_fr_costs = wb.create_sheet("France Costs History")
        exercises = [e for e in fr_data["exercises"] if e.get('date_publication')]
        
        cols = [
            ("Year", "annee_debut", 8),
            ("Lobbying Costs", "montant_depense", 30),
            ("Min €", "montant_depense_inf", 15),
            ("Max €", "montant_depense_sup", 15),
            ("Staff", "nombre_salaries", 10),
            ("Activities", "nombre_activites", 10),
        ]
        
        ws_fr_costs['A1'] = f"French Lobbying Costs Over Time - {org_name}"
        ws_fr_costs['A1'].font = Font(bold=True, size=14)
        
        for col_idx, (header, _, width) in enumerate(cols, 1):
            cell = ws_fr_costs.cell(row=3, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill_fr
            ws_fr_costs.column_dimensions[get_column_letter(col_idx)].width = width
        
        for row_idx, e in enumerate(exercises, 4):
            for col_idx, (_, field, _) in enumerate(cols, 1):
                value = e.get(field, "")
                if field in ("montant_depense_inf", "montant_depense_sup") and value:
                    try:
                        value = int(float(value))
                    except:
                        pass
                ws_fr_costs.cell(row=row_idx, column=col_idx, value=value)
    
    # === FRANCE OFFICIALS TARGETED ===
    if fr_data and fr_data.get("officials"):
        ws_officials = wb.create_sheet("France Officials Targeted")
        
        # Aggregate
        official_counts = {}
        ministry_counts = {}
        for o in fr_data["officials"]:
            off = o.get('responsable_public', 'Unknown')
            min_ = o.get('departement_ministeriel', '')
            official_counts[off] = official_counts.get(off, 0) + 1
            if min_:
                ministry_counts[min_] = ministry_counts.get(min_, 0) + 1
        
        ws_officials['A1'] = f"French Officials/Institutions Targeted - {org_name}"
        ws_officials['A1'].font = Font(bold=True, size=14)
        
        ws_officials['A3'] = "By Official Type"
        ws_officials['A3'].font = Font(bold=True)
        ws_officials['A4'] = "Type"
        ws_officials['B4'] = "Count"
        ws_officials['A4'].font = header_font
        ws_officials['B4'].font = header_font
        ws_officials['A4'].fill = header_fill_fr
        ws_officials['B4'].fill = header_fill_fr
        
        for i, (off, count) in enumerate(sorted(official_counts.items(), key=lambda x: -x[1]), 5):
            ws_officials.cell(row=i, column=1, value=off[:80])
            ws_officials.cell(row=i, column=2, value=count)
        
        start_row = 5 + len(official_counts) + 2
        ws_officials.cell(row=start_row, column=1, value="By Ministry/Department").font = Font(bold=True)
        ws_officials.cell(row=start_row+1, column=1, value="Ministry").font = header_font
        ws_officials.cell(row=start_row+1, column=2, value="Count").font = header_font
        ws_officials.cell(row=start_row+1, column=1).fill = header_fill_fr
        ws_officials.cell(row=start_row+1, column=2).fill = header_fill_fr
        
        for i, (min_, count) in enumerate(sorted(ministry_counts.items(), key=lambda x: -x[1]), start_row+2):
            ws_officials.cell(row=i, column=1, value=min_)
            ws_officials.cell(row=i, column=2, value=count)
        
        ws_officials.column_dimensions['A'].width = 80
        ws_officials.column_dimensions['B'].width = 10
    
    # === GERMANY REGULATORY PROJECTS SHEET ===
    if de_data and de_data.get("regulatory_projects"):
        ws_de_proj = wb.create_sheet("Germany Regulatory Projects")
        
        ws_de_proj['A1'] = f"German Regulatory Projects Lobbied - {org_name}"
        ws_de_proj['A1'].font = Font(bold=True, size=14)
        ws_de_proj.merge_cells('A1:D1')
        
        ws_de_proj['A2'] = "Source: Bundestag Lobbyregister"
        ws_de_proj['A2'].font = Font(italic=True, size=10)
        
        cols = [("Project Title", 60), ("Ministry", 15), ("Document", 20), ("URL", 50)]
        for col_idx, (header, width) in enumerate(cols, 1):
            cell = ws_de_proj.cell(row=4, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill_de
            ws_de_proj.column_dimensions[get_column_letter(col_idx)].width = width
        
        row_idx = 5
        for proj in de_data["regulatory_projects"]:
            for pm in proj.get("printed_matters", []):
                ws_de_proj.cell(row=row_idx, column=1, value=proj.get("title", "")[:500])
                ws_de_proj.cell(row=row_idx, column=2, value=pm.get("ministry", ""))
                ws_de_proj.cell(row=row_idx, column=3, value=pm.get("number", ""))
                ws_de_proj.cell(row=row_idx, column=4, value=pm.get("url", ""))
                row_idx += 1
        
        ws_de_proj.freeze_panes = 'A5'
    
    # === GERMANY FIELDS OF INTEREST SHEET ===
    if de_data and de_data.get("fields_of_interest"):
        ws_de_fields = wb.create_sheet("Germany Fields of Interest")
        
        ws_de_fields['A1'] = f"German Lobbying Fields of Interest - {org_name}"
        ws_de_fields['A1'].font = Font(bold=True, size=14)
        
        ws_de_fields.cell(row=3, column=1, value="Field of Interest").font = header_font
        ws_de_fields.cell(row=3, column=1).fill = header_fill_de
        ws_de_fields.column_dimensions['A'].width = 60
        
        for i, field in enumerate(de_data["fields_of_interest"], 4):
            ws_de_fields.cell(row=i, column=1, value=field)
    
    # === IRELAND LOBBYING RETURNS SHEET ===
    if ie_data and ie_data.get("returns"):
        ws_ie = wb.create_sheet("Ireland Lobbying Returns")
        
        ws_ie['A1'] = f"Irish Lobbying Returns - {org_name}"
        ws_ie['A1'].font = Font(bold=True, size=14)
        ws_ie.merge_cells('A1:F1')
        
        ws_ie['A2'] = "Source: lobbying.ie | Note: Ireland does not track lobbying expenditure"
        ws_ie['A2'].font = Font(italic=True, size=10)
        
        cols = [
            ("Return Period", 20), ("Lobbyist", 35), ("Client", 25),
            ("Subject Matter", 30), ("Policy Area", 25),
            ("Public Body", 30), ("Official", 30), ("Job Title", 20),
            ("Subject Details", 50), ("Submitted", 15)
        ]
        for col_idx, (header, width) in enumerate(cols, 1):
            cell = ws_ie.cell(row=4, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill_ie
            ws_ie.column_dimensions[get_column_letter(col_idx)].width = width
        
        for row_idx, ret in enumerate(ie_data["returns"], 5):
            ws_ie.cell(row=row_idx, column=1, value=ret.get("return_period", ""))
            ws_ie.cell(row=row_idx, column=2, value=ret.get("lobbyist", ""))
            ws_ie.cell(row=row_idx, column=3, value=ret.get("client", ""))
            ws_ie.cell(row=row_idx, column=4, value=ret.get("subject_matter", ""))
            ws_ie.cell(row=row_idx, column=5, value=ret.get("subject_matter_area", ""))
            ws_ie.cell(row=row_idx, column=6, value=ret.get("public_body", ""))
            ws_ie.cell(row=row_idx, column=7, value=ret.get("dpo_name", ""))
            ws_ie.cell(row=row_idx, column=8, value=ret.get("dpo_title", ""))
            ws_ie.cell(row=row_idx, column=9, value=ret.get("subject_matter_details", "")[:500] if ret.get("subject_matter_details") else "")
            ws_ie.cell(row=row_idx, column=10, value=ret.get("submitted_date", ""))
        
        ws_ie.freeze_panes = 'A5'
    
    # === UK MINISTERIAL MEETINGS SHEET ===
    if uk_data and uk_data.get("meetings"):
        ws_uk = wb.create_sheet("UK Ministerial Meetings")
        
        ws_uk['A1'] = f"UK Ministerial Meetings - {org_name}"
        ws_uk['A1'].font = Font(bold=True, size=14)
        ws_uk.merge_cells('A1:E1')
        
        ws_uk['A2'] = "Source: GOV.UK Transparency Publications | Note: UK does not track lobbying expenditure"
        ws_uk['A2'].font = Font(italic=True, size=10)
        
        cols = [
            ("Date", 12), ("Minister", 30), ("Organisation", 40),
            ("Purpose", 50), ("Department", 30)
        ]
        for col_idx, (header, width) in enumerate(cols, 1):
            cell = ws_uk.cell(row=4, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill_uk
            ws_uk.column_dimensions[get_column_letter(col_idx)].width = width
        
        for row_idx, meeting in enumerate(uk_data["meetings"][:1000], 5):  # Limit to 1000 rows
            ws_uk.cell(row=row_idx, column=1, value=meeting.get("date", ""))
            ws_uk.cell(row=row_idx, column=2, value=meeting.get("minister", ""))
            ws_uk.cell(row=row_idx, column=3, value=meeting.get("organisation", ""))
            ws_uk.cell(row=row_idx, column=4, value=meeting.get("purpose", ""))
            ws_uk.cell(row=row_idx, column=5, value=meeting.get("department", ""))
        
        ws_uk.freeze_panes = 'A5'
        
        # Add UK by-minister summary sheet
        if uk_data.get("by_minister"):
            ws_uk_min = wb.create_sheet("UK By Minister")
            ws_uk_min['A1'] = f"UK Meetings by Minister - {org_name}"
            ws_uk_min['A1'].font = Font(bold=True, size=14)
            
            ws_uk_min.cell(row=3, column=1, value="Minister").font = header_font
            ws_uk_min.cell(row=3, column=2, value="Meetings").font = header_font
            ws_uk_min.cell(row=3, column=1).fill = header_fill_uk
            ws_uk_min.cell(row=3, column=2).fill = header_fill_uk
            
            for row_idx, (minister, count) in enumerate(uk_data["by_minister"].items(), 4):
                ws_uk_min.cell(row=row_idx, column=1, value=minister)
                ws_uk_min.cell(row=row_idx, column=2, value=count)
            
            ws_uk_min.column_dimensions['A'].width = 40
            ws_uk_min.column_dimensions['B'].width = 15
    
    # === UK SENIOR OFFICIALS MEETINGS SHEET ===
    if uk_officials_data and uk_officials_data.get("meetings"):
        ws_uk_so = wb.create_sheet("UK Senior Officials")
        
        ws_uk_so['A1'] = f"UK Senior Officials Meetings - {org_name}"
        ws_uk_so['A1'].font = Font(bold=True, size=14)
        ws_uk_so.merge_cells('A1:E1')
        
        ws_uk_so['A2'] = "Source: GOV.UK Transparency | Officials: Permanent Secretaries, DGs, SCS2+"
        ws_uk_so['A2'].font = Font(italic=True, size=10)
        
        cols = [
            ("Date", 12), ("Official", 30), ("Organisation", 40),
            ("Purpose", 50), ("Department", 30)
        ]
        for col_idx, (header, width) in enumerate(cols, 1):
            cell = ws_uk_so.cell(row=4, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill_uk_officials
            ws_uk_so.column_dimensions[get_column_letter(col_idx)].width = width
        
        for row_idx, meeting in enumerate(uk_officials_data["meetings"][:1000], 5):
            ws_uk_so.cell(row=row_idx, column=1, value=meeting.get("date", ""))
            ws_uk_so.cell(row=row_idx, column=2, value=meeting.get("minister", ""))  # Field is still called minister
            ws_uk_so.cell(row=row_idx, column=3, value=meeting.get("organisation", ""))
            ws_uk_so.cell(row=row_idx, column=4, value=meeting.get("purpose", ""))
            ws_uk_so.cell(row=row_idx, column=5, value=meeting.get("department", ""))
        
        ws_uk_so.freeze_panes = 'A5'
        
        # Add by-official summary sheet
        if uk_officials_data.get("by_official"):
            ws_uk_off = wb.create_sheet("UK By Official")
            ws_uk_off['A1'] = f"UK Meetings by Senior Official - {org_name}"
            ws_uk_off['A1'].font = Font(bold=True, size=14)
            
            ws_uk_off.cell(row=3, column=1, value="Official").font = header_font
            ws_uk_off.cell(row=3, column=2, value="Meetings").font = header_font
            ws_uk_off.cell(row=3, column=1).fill = header_fill_uk_officials
            ws_uk_off.cell(row=3, column=2).fill = header_fill_uk_officials
            
            for row_idx, (official, count) in enumerate(uk_officials_data["by_official"].items(), 4):
                ws_uk_off.cell(row=row_idx, column=1, value=official)
                ws_uk_off.cell(row=row_idx, column=2, value=count)
            
            ws_uk_off.column_dimensions['A'].width = 40
            ws_uk_off.column_dimensions['B'].width = 15
    
    # === AUSTRIA REGISTER SHEET ===
    if at_data and at_data.get("entries"):
        ws_at = wb.create_sheet("Austria Register")
        
        ws_at['A1'] = f"Austrian Lobbying Register - {org_name}"
        ws_at['A1'].font = Font(bold=True, size=14)
        ws_at.merge_cells('A1:F1')
        
        ws_at['A2'] = "Source: lobbyreg.justiz.gv.at | Note: Costs only disclosed if >€100,000"
        ws_at['A2'].font = Font(italic=True, size=10)
        
        cols = [
            ("Organisation", 50), ("Register #", 15), ("Category", 12),
            ("Category Desc", 25), ("Lobbyists", 40), ("Last Update", 12)
        ]
        for col_idx, (header, width) in enumerate(cols, 1):
            cell = ws_at.cell(row=4, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill_at
            ws_at.column_dimensions[get_column_letter(col_idx)].width = width
        
        for row_idx, entry in enumerate(at_data["entries"][:500], 5):  # Limit to 500 rows
            ws_at.cell(row=row_idx, column=1, value=entry.get("name", ""))
            ws_at.cell(row=row_idx, column=2, value=entry.get("register_number", ""))
            ws_at.cell(row=row_idx, column=3, value=entry.get("category", ""))
            ws_at.cell(row=row_idx, column=4, value=entry.get("category_description", ""))
            ws_at.cell(row=row_idx, column=5, value=entry.get("lobbyists", ""))
            ws_at.cell(row=row_idx, column=6, value=entry.get("last_update", ""))
        
        ws_at.freeze_panes = 'A5'
    
    # === CATALONIA REGISTER SHEET ===
    if cat_data and cat_data.get("entries"):
        ws_cat = wb.create_sheet("Catalonia Register")
        
        ws_cat['A1'] = f"Catalonia Interest Group Register - {org_name}"
        ws_cat['A1'].font = Font(bold=True, size=14)
        ws_cat.merge_cells('A1:H1')
        
        ws_cat['A2'] = "Source: analisi.transparenciacatalunya.cat | Registre de grups d'interès de Catalunya"
        ws_cat['A2'].font = Font(italic=True, size=10)
        
        cols = [
            ("Organisation", 40), ("ID", 10), ("Reg. Date", 12),
            ("Category", 35), ("Annual Volume", 15), ("Areas of Interest", 40),
            ("Purpose", 50), ("Province", 15)
        ]
        for col_idx, (header, width) in enumerate(cols, 1):
            cell = ws_cat.cell(row=4, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill_cat
            cell.font = Font(bold=True, color="000000")  # Black text on yellow
            ws_cat.column_dimensions[get_column_letter(col_idx)].width = width
        
        for row_idx, entry in enumerate(cat_data["entries"][:500], 5):  # Limit to 500 rows
            ws_cat.cell(row=row_idx, column=1, value=entry.get("name", ""))
            ws_cat.cell(row=row_idx, column=2, value=entry.get("id", ""))
            ws_cat.cell(row=row_idx, column=3, value=entry.get("registration_date", ""))
            ws_cat.cell(row=row_idx, column=4, value=entry.get("category", ""))
            ws_cat.cell(row=row_idx, column=5, value=entry.get("annual_volume_formatted", ""))
            ws_cat.cell(row=row_idx, column=6, value=entry.get("areas_of_interest", "")[:100] if entry.get("areas_of_interest") else "")
            ws_cat.cell(row=row_idx, column=7, value=entry.get("purpose", "")[:200] if entry.get("purpose") else "")
            ws_cat.cell(row=row_idx, column=8, value=entry.get("province", ""))
        
        ws_cat.freeze_panes = 'A5'
    
    # === FINLAND REGISTER SHEET ===
    if fi_data and fi_data.get("entries"):
        ws_fi = wb.create_sheet("Finland Register")
        
        ws_fi['A1'] = f"Finland Transparency Register - {org_name}"
        ws_fi['A1'].font = Font(bold=True, size=14)
        ws_fi.merge_cells('A1:G1')
        
        ws_fi['A2'] = "Source: avoimuusrekisteri.fi | Financial data available from July 2026"
        ws_fi['A2'].font = Font(italic=True, size=10)
        
        cols = [
            ("Organisation", 40), ("Company ID", 15), ("Diary #", 18),
            ("Reg. Date", 12), ("Industry", 35), ("Activities", 12), ("Topics", 50)
        ]
        for col_idx, (header, width) in enumerate(cols, 1):
            cell = ws_fi.cell(row=4, column=col_idx, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = header_fill_fi
            ws_fi.column_dimensions[get_column_letter(col_idx)].width = width
        
        for row_idx, entry in enumerate(fi_data["entries"][:500], 5):
            ws_fi.cell(row=row_idx, column=1, value=entry.get("name", ""))
            ws_fi.cell(row=row_idx, column=2, value=entry.get("company_id", ""))
            ws_fi.cell(row=row_idx, column=3, value=entry.get("diary_number", ""))
            ws_fi.cell(row=row_idx, column=4, value=entry.get("registration_date", ""))
            ws_fi.cell(row=row_idx, column=5, value=entry.get("main_industry", ""))
            ws_fi.cell(row=row_idx, column=6, value=entry.get("activity_count", 0))
            topics_str = "; ".join(entry.get("topics", []))[:200]
            ws_fi.cell(row=row_idx, column=7, value=topics_str)
        
        ws_fi.freeze_panes = 'A5'
    
    # === SLOVENIA REGISTER SHEET ===
    if si_data and si_data.get("entries"):
        ws_si = wb.create_sheet("Slovenia Lobbyists")
        
        ws_si['A1'] = f"Slovenia Register of Lobbyists - {org_name}"
        ws_si['A1'].font = Font(bold=True, size=14)
        ws_si.merge_cells('A1:F1')
        
        ws_si['A2'] = "Source: kpk-rs.si | Lists individual lobbyists, not companies"
        ws_si['A2'].font = Font(italic=True, size=10)
        
        cols = [
            ("Lobbyist Name", 30), ("Company/Employer", 35), ("Fields of Interest", 60),
            ("Address", 30), ("City", 20), ("Email", 30)
        ]
        for col_idx, (header, width) in enumerate(cols, 1):
            cell = ws_si.cell(row=4, column=col_idx, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = header_fill_si
            ws_si.column_dimensions[get_column_letter(col_idx)].width = width
        
        for row_idx, entry in enumerate(si_data["entries"][:500], 5):
            ws_si.cell(row=row_idx, column=1, value=entry.get("name", ""))
            ws_si.cell(row=row_idx, column=2, value=entry.get("company", ""))
            fields_str = " · ".join(entry.get("fields_of_interest", []))[:300]
            ws_si.cell(row=row_idx, column=3, value=fields_str)
            ws_si.cell(row=row_idx, column=4, value=entry.get("address", ""))
            ws_si.cell(row=row_idx, column=5, value=entry.get("city", ""))
            ws_si.cell(row=row_idx, column=6, value=entry.get("email", ""))
        
        ws_si.freeze_panes = 'A5'
    
    wb.save(output_path)
    print(f"\nReport saved to: {output_path}")


def print_summary(eu_data: dict, fr_data: dict, de_data: dict, ie_data: dict, uk_data: dict, at_data: dict, cat_data: dict, fi_data: dict, org_name: str):
    """Print summary to console."""
    print(f"\n{'='*60}")
    print(f"Summary for: {org_name}")
    print(f"{'='*60}")
    
    if eu_data:
        regs = eu_data.get("registrations", [])
        meetings = eu_data.get("meetings", [])
        latest = regs[-1] if regs else {}
        
        print(f"\n[EU - LobbyFacts]")
        print(f"  Lobbying costs: €{int(latest.get('min', 0) or 0):,} - €{int(latest.get('max', 0) or 0):,}" if latest.get('min') or latest.get('max') else "  Lobbying costs: Not disclosed")
        print(f"  Commission meetings: {len(meetings)}")
        
        if meetings:
            by_year = {}
            for m in meetings:
                year = m.get("Date", "")[:4]
                if year:
                    by_year[year] = by_year.get(year, 0) + 1
            print("  Meetings by year:")
            for year in sorted(by_year.keys(), reverse=True)[:5]:
                print(f"    {year}: {by_year[year]}")
    
    if fr_data and fr_data.get("info"):
        exercises = fr_data.get("exercises", [])
        activities = fr_data.get("activities", [])
        latest = exercises[2] if len(exercises) > 2 else (exercises[0] if exercises else {})
        
        print(f"\n[France - HATVP]")
        print(f"  Lobbying costs: {latest.get('montant_depense', 'Not disclosed')}")
        print(f"  Declared activities: {len(activities)}")
        
        if activities:
            by_year = {}
            for a in activities:
                year = a.get("date_publication_activite", "")[:4]
                if year:
                    by_year[year] = by_year.get(year, 0) + 1
            print("  Activities by year:")
            for year in sorted(by_year.keys(), reverse=True)[:5]:
                print(f"    {year}: {by_year[year]}")
    
    if de_data and de_data.get("name"):
        print(f"\n[Germany - Bundestag Lobbyregister]")
        print(f"  Lobbying costs: €{de_data.get('expenses_min', 0):,} - €{de_data.get('expenses_max', 0):,}")
        print(f"  Employee FTE: {de_data.get('employee_fte', 'N/A')}")
        print(f"  Regulatory projects: {de_data.get('regulatory_projects_count', 0)}")
        print(f"  Fields of interest: {len(de_data.get('fields_of_interest', []))}")
    
    if ie_data and ie_data.get("name"):
        print(f"\n[Ireland - Lobbying.ie]")
        print(f"  Lobbying costs: Not tracked (not required under Irish law)")
        print(f"  Lobbying returns: {ie_data.get('returns_count', 0)}")
        if ie_data.get("returns"):
            # Count by period
            by_period = {}
            for ret in ie_data["returns"]:
                period = ret.get("return_period", "Unknown")
                by_period[period] = by_period.get(period, 0) + 1
            print("  Returns by period (recent):")
            for period in sorted(by_period.keys(), reverse=True)[:5]:
                print(f"    {period}: {by_period[period]}")
    
    if uk_data and uk_data.get("meetings"):
        print(f"\n[UK - GOV.UK Ministerial Meetings]")
        print(f"  Lobbying costs: Not tracked (UK publishes meetings only)")
        print(f"  Ministerial meetings: {uk_data.get('meetings_count', 0)}")
        if uk_data.get("by_year"):
            print("  Meetings by year:")
            for year, count in list(uk_data["by_year"].items())[:5]:
                print(f"    {year}: {count}")
        if uk_data.get("by_minister"):
            print("  Top ministers met:")
            for minister, count in list(uk_data["by_minister"].items())[:3]:
                print(f"    {minister}: {count}")
    
    if at_data and at_data.get("entries"):
        print(f"\n[Austria - Lobbying Register]")
        print(f"  Lobbying costs: Only disclosed if >€100,000")
        print(f"  Register entries: {at_data.get('entry_count', 0)}")
        if at_data.get("by_category"):
            print("  By category:")
            for cat, count in at_data["by_category"].items():
                cat_desc = {"A1": "Lobbying firm", "B": "In-house", "C": "Self-gov body", "D": "Interest group"}.get(cat, cat)
                print(f"    {cat} ({cat_desc}): {count}")
    
    if cat_data and cat_data.get("entries"):
        print(f"\n[Catalonia - Interest Group Register]")
        print(f"  Register entries: {cat_data.get('entry_count', 0)}")
        print(f"  Total annual volume: {cat_data.get('total_volume_formatted', 'N/A')}")
        if cat_data.get("by_category"):
            print("  By category:")
            for cat, count in cat_data["by_category"].items():
                print(f"    {cat}: {count}")
    
    if fi_data and fi_data.get("entries"):
        print(f"\n[Finland - Transparency Register]")
        print(f"  Register entries: {fi_data.get('entry_count', 0)}")
        print(f"  Activity disclosures: {fi_data.get('total_activities', 0)}")
        print(f"  Lobbying costs: Financial data available from July 2026")
        for entry in fi_data.get("entries", [])[:3]:
            if entry.get("topics"):
                print(f"  Topics: {'; '.join(entry['topics'][:3])}")


def main():
    parser = argparse.ArgumentParser(
        description="European Lobbying Data Tool - EU + France",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
    python eu_lobbying.py --company "Google"
    python eu_lobbying.py --company "Microsoft" --output ms_report.xlsx
        """
    )
    
    parser.add_argument("--company", "-c", help="Company name to search for")
    parser.add_argument("--eu-id", help="Direct EU Transparency Register ID")
    parser.add_argument("--fr-id", help="Direct French HATVP ID")
    parser.add_argument("--de-reg", help="Direct German Lobbyregister number (e.g., R001794)")
    parser.add_argument("--ie-csv", help="Path to Irish lobbying.ie CSV export (manual download required)")
    parser.add_argument("--ie-help", action="store_true", help="Show instructions for downloading Irish lobbying data")
    parser.add_argument("--skip-uk", action="store_true", help="Skip UK ministerial meetings search (faster)")
    parser.add_argument("--output", "-o", help="Output Excel file path")
    
    args = parser.parse_args()
    
    # Show Ireland download instructions if requested
    if args.ie_help:
        print_ireland_instructions()
        sys.exit(0)
    
    if not args.company and not args.eu_id and not args.fr_id and not args.de_reg and not args.ie_csv:
        parser.print_help()
        sys.exit(1)
    
    eu_id = args.eu_id
    fr_id = args.fr_id
    de_reg = args.de_reg
    ie_csv = args.ie_csv
    skip_uk = args.skip_uk
    org_name = args.company or "Organisation"
    
    # Search if company name provided
    if args.company:
        # Search EU
        print(f"\nSearching EU register for '{args.company}'...")
        eu_results = search_eu_register(args.company)
        if eu_results:
            if len(eu_results) == 1:
                eu_id = eu_results[0]["id"]
                org_name = eu_results[0]["name"]
                print(f"  Found: {org_name} (ID: {eu_id})")
            else:
                print(f"  Found {len(eu_results)} matches:")
                for r in eu_results[:5]:
                    print(f"    - {r['name']} (ID: {r['id']})")
                if len(eu_results) == 1:
                    eu_id = eu_results[0]["id"]
        else:
            print("  No EU matches found")
        
        # Search France
        print(f"\nSearching French register for '{args.company}'...")
        fr_results = search_france_register(args.company)
        if fr_results:
            if len(fr_results) == 1:
                fr_id = fr_results[0]["id"]
                print(f"  Found: {fr_results[0]['name']} (ID: {fr_id})")
            else:
                print(f"  Found {len(fr_results)} matches:")
                for r in fr_results[:5]:
                    print(f"    - {r['name']} (ID: {r['id']})")
                if len(fr_results) == 1:
                    fr_id = fr_results[0]["id"]
        else:
            print("  No French matches found")
        
        # Search Germany
        print(f"\nSearching German register for '{args.company}'...")
        de_results = search_germany_register(args.company)
        if de_results:
            # Auto-select if first result contains company name
            search_lower = args.company.lower()
            first_name_lower = de_results[0]["name"].lower()
            if len(de_results) == 1 or search_lower in first_name_lower:
                de_reg = de_results[0]["register_number"]
                print(f"  Found: {de_results[0]['name']} (Reg: {de_reg})")
            else:
                print(f"  Found {len(de_results)} matches:")
                for r in de_results[:5]:
                    print(f"    - {r['name']} (Reg: {r['register_number']})")
                print(f"  Use --de-reg to specify (e.g., --de-reg {de_results[0]['register_number']})")
        else:
            print("  No German matches found")
        
        # Print Ireland instructions (no API available)
        print(f"\n[Ireland - Manual Download Required]")
        print(f"  The Irish register doesn't have a public API.")
        print(f"  To include Irish data, use: --ie-csv path/to/downloaded.csv")
        print(f"  For instructions, use: --ie-help")
    
    # Fetch data
    eu_data = None
    fr_data = None
    de_data = None
    ie_data = None
    uk_data = None
    
    if eu_id:
        try:
            eu_data = fetch_eu_data(eu_id)
        except Exception as e:
            print(f"Error fetching EU data: {e}")
    
    if fr_id:
        try:
            fr_data = fetch_france_data(fr_id)
        except Exception as e:
            print(f"Error fetching French data: {e}")
    
    if de_reg:
        try:
            de_data = fetch_germany_data(de_reg)
        except Exception as e:
            print(f"Error fetching German data: {e}")
    
    if ie_csv:
        try:
            ie_data = load_ireland_csv(ie_csv, args.company)
        except Exception as e:
            print(f"Error loading Irish data: {e}")
    
    # Search UK ministerial meetings (unless skipped)
    if args.company and not skip_uk:
        try:
            uk_data = search_uk_ministerial_meetings(args.company)
        except Exception as e:
            print(f"Error fetching UK data: {e}")
    
    # Search Austria register
    at_data = None
    if args.company:
        try:
            at_data = search_austria_register(args.company)
        except Exception as e:
            print(f"Error fetching Austrian data: {e}")
    
    # Search Catalonia register
    cat_data = None
    if args.company:
        try:
            cat_data = search_catalonia_register(args.company)
        except Exception as e:
            print(f"Error fetching Catalonia data: {e}")
    
    # Search Finland register
    fi_data = None
    if args.company:
        try:
            fi_data = search_finland_register(args.company)
        except Exception as e:
            print(f"Error fetching Finland data: {e}")
    
    if not eu_data and not fr_data and not de_data and not ie_data and not uk_data and not at_data and not cat_data and not fi_data:
        print("\nNo data found. Try specifying IDs directly with --eu-id, --fr-id, --de-reg, or --ie-csv")
        sys.exit(1)
    
    # Generate output
    safe_name = re.sub(r'[^\w\-]', '_', org_name.lower())
    output_file = args.output or f"{safe_name}_lobbying.xlsx"
    
    create_excel_report(eu_data, fr_data, de_data, ie_data, uk_data, at_data, cat_data, fi_data, output_file, org_name)
    print_summary(eu_data, fr_data, de_data, ie_data, uk_data, at_data, cat_data, fi_data, org_name)


if __name__ == "__main__":
    main()
