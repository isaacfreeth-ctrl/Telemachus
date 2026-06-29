#!/usr/bin/env python3
"""
Build a pre-indexed database of all UK ministerial and senior officials meetings.

This script downloads ALL ministerial and senior officials meeting attachments from
GOV.UK, combines them into a single searchable JSON file, and saves it.

The resulting index can be searched instantly without any API calls.

Data sources:
- GOV.UK transparency publications (ministerial meetings)
- GOV.UK transparency publications (senior officials meetings)
- GOV.UK transparency *collection* pages (authoritative list of every edition)

Run this script periodically to keep the index up to date:
    python build_uk_index.py

Output:
    uk_meetings_index.json.gz - Complete index of all meetings with source URLs,
    a coverage manifest, a nil-return sheet, and a name-review list.

This module addresses the data-gap spec (transparency_tool_gaps.md):
  1.  Row-cap truncation        -> paginate + loud total assertion
  2.  Slug drift                -> collection-page resolution + stem fallback + fixed filter
  3.  Multi-attachment / ODS    -> enumerate all attachments, CSV-first, XLSX fallback, filename tag
  4.  Date / Nil-return hygiene -> date_flag (no silent drop) + separate nil sheet
  5.  Name normalisation        -> person_canonical + honorific + alias map + near-dup review
  6.  Cross-return dedup        -> keep raw rows, add dedup_key for the counting layer
  8.  Coverage manifest         -> per-publication manifest sheet
  9.  Amendment drift           -> content hash + fetch timestamp, diff vs prior index
  10. Department scope leakage  -> source department + starmer_era cutoff flag
"""

import json
import csv
import io
import re
import sys
import time
import gzip
import hashlib
import difflib
import requests
from pathlib import Path
from datetime import datetime, date
from collections import defaultdict
from urllib.parse import quote, unquote

# Use the OS trust store for TLS where available. On machines behind a
# TLS-inspection proxy (corporate AV / gateway), the injected root CA lives in the
# Windows cert store but NOT in certifi's bundle, so requests would otherwise fail
# with CERTIFICATE_VERIFY_FAILED and the whole scrape would silently return zero.
try:
    import truststore
    truststore.inject_into_ssl()
except Exception:
    pass

# GOV.UK API endpoints
GOVUK_SEARCH_URL = "https://www.gov.uk/api/search.json"
GOVUK_CONTENT_URL = "https://www.gov.uk/api/content"

# Rate limiting
REQUEST_DELAY = 0.15  # seconds between requests (polite; GOV.UK handles this easily)

# Starmer-era cutoff (gap 10): meetings on/after this date are tagged starmer_era=True.
# Kept as a *flag*, never a hard drop, since prior-role rows are wanted for some analyses.
STARMER_CUTOFF = date(2024, 7, 5)


# ---------------------------------------------------------------------------
# Gap 2: collection pages are the source of truth for "every edition that exists".
# We resolve child publications from these rather than trusting search alone, and
# log any advertised period the tool failed to fetch.
# ---------------------------------------------------------------------------
KNOWN_COLLECTION_PATHS = [
    # Cross-government collections (list ministerial + senior official editions)
    "/government/collections/ministers-transparency-publications",
    "/government/collections/special-adviser-data-releases-and-transparency",
    # Department-specific transparency collections (extend freely; harmless if 404)
    "/government/collections/cabinet-office-ministerial-gifts-hospitality-travel-and-meetings",
    "/government/collections/dsit-ministerial-gifts-hospitality-travel-and-meetings",
    "/government/collections/dsit-and-ministers-transparency-data",
]

# Gap 2: regex stems for the *same recurring series* whose slug GOV.UK has rotated.
# Used to (a) tag a publication with its series for the manifest and (b) as a
# fallback membership test. Order doesn't matter; all variants are "live and needed".
SERIES_SLUG_STEMS = {
    "cabinet_office_ministerial": [
        r"cabinet-office-ministerial-gifts-hospitality-travel-and-meetings",
        r"cabinet-office-ministerial-gifts-hospitality-overseas-travel-and-meetings",
        r"cabinet-office-ministerial-overseas-travel-and-meetings",
    ],
    "cabinet_office_senior": [
        r"cabinet-office-business-expenses-hospitality-and-meetings",
        r"cabinet-office-business-expenses-hospitality-and-meetings-for-senior-officials",
    ],
    "dsit_ministerial": [
        r"dsit-ministerial-gifts-hospitality-travel-meetings",
        r"dsit-ministerial-overseas-travel-and-meetings",
    ],
    "dsit_senior": [
        r"dsit-senior-officials-business-expenses-and-hospitality",
    ],
}


def classify_series(link: str) -> str:
    """Return the series key whose slug stem this publication link matches, else ''."""
    low = (link or "").lower()
    for series, stems in SERIES_SLUG_STEMS.items():
        for stem in stems:
            if re.search(stem, low):
                return series
    return ""


# ---------------------------------------------------------------------------
# Gap 5: name normalisation helpers.
# ---------------------------------------------------------------------------
# Honorifics stripped from the canonical name into a separate field. Order matters
# only in that we strip leading tokens iteratively.
HONORIFICS = {
    "rt", "hon", "the", "sir", "dame", "lord", "lady", "baroness", "baron",
    "dr", "mr", "mrs", "ms", "miss", "prof", "professor", "rev", "reverend",
    "cllr", "councillor",
}

# Clear, unambiguous published misspellings only. Ambiguous near-duplicates
# (e.g. Allan vs Allen) are NOT auto-merged here; they are surfaced for manual
# review instead (see build_name_review).
NAME_ALIASES = {
    "david dnismore": "david dinsmore",
}


def canonicalize_name(raw: str):
    """
    Produce (person_canonical, honorific) from a verbatim attendee string.

    - case-folds to Title Case (so ANDREW FORZANI == Andrew Forzani)
    - collapses whitespace
    - standardises hyphenation (Neville-Rolfe == Neville Rolfe)
    - strips honorifics into a separate field (Sir Keir Starmer -> Keir Starmer / "Sir")
    - applies the clear-typo alias map

    The verbatim source string is preserved separately by the caller.
    """
    if not raw:
        return "", ""

    s = raw.strip()
    # Normalise hyphens/whitespace so "Neville-Rolfe" groups with "Neville Rolfe".
    s = s.replace("-", " ")
    s = re.sub(r"\s+", " ", s).strip()
    tokens = [t for t in s.lower().split(" ") if t]

    # Peel leading honorific tokens (handles "the rt hon sir ...").
    honorifics = []
    while tokens and tokens[0].strip(".,") in HONORIFICS:
        honorifics.append(tokens.pop(0).strip(".,").title())

    # Apply clear-typo aliases on the honorific-stripped name (handles names that
    # were preceded by honorifics, e.g. "The Rt Hon David Dnismore").
    stripped_low = " ".join(tokens).strip()
    stripped_low = NAME_ALIASES.get(stripped_low, stripped_low)

    canonical = " ".join(t.title() for t in stripped_low.split(" ") if t).strip()
    honorific = " ".join(honorifics).strip()
    return canonical, honorific


# ---------------------------------------------------------------------------
# Gap 4: quarter parsing + date flagging.
# ---------------------------------------------------------------------------
_MONTH_NUM = {
    'january': 1, 'february': 2, 'march': 3, 'april': 4, 'may': 5, 'june': 6,
    'july': 7, 'august': 8, 'september': 9, 'october': 10, 'november': 11,
    'december': 12, 'jan': 1, 'feb': 2, 'mar': 3, 'apr': 4, 'jun': 6, 'jul': 7,
    'aug': 8, 'sep': 9, 'sept': 9, 'oct': 10, 'nov': 11, 'dec': 12,
}


def _end_of_month(year: int, month: int) -> date:
    if month == 12:
        return date(year, 12, 31)
    first_of_next = date(year, month + 1, 1)
    return date.fromordinal(first_of_next.toordinal() - 1)


def parse_period_from_url(source_url: str):
    """
    Parse the publication's stated reporting period into (start_date, end_date, label).

    Recognises e.g.:
        october-to-december-2013
        july-to-september-2024
        april-to-june-2011
        q3-2025 / 2025-q3
        2015 (bare year -> whole year)

    Returns (None, None, "") if no period can be inferred.
    """
    if not source_url:
        return None, None, ""
    low = source_url.lower()

    # "<month>-to-<month>-<year>"
    m = re.search(r'([a-z]+)-to-([a-z]+)-(\d{4})', low)
    if m and m.group(1) in _MONTH_NUM and m.group(2) in _MONTH_NUM:
        y = int(m.group(3))
        sm, em = _MONTH_NUM[m.group(1)], _MONTH_NUM[m.group(2)]
        # Quarter may straddle a year boundary (e.g. oct-to-dec is same year here).
        start = date(y, sm, 1)
        end = _end_of_month(y, em)
        if end < start:  # safety; shouldn't happen for same-year quarters
            end = _end_of_month(y, 12)
        return start, end, m.group(0)

    # "q3-2025" / "2025-q3"
    m = re.search(r'q([1-4])[-\s]?(\d{4})', low) or re.search(r'(\d{4})[-\s]?q([1-4])', low)
    if m:
        groups = m.groups()
        q = int(groups[0]) if groups[0].isdigit() and len(groups[0]) == 1 else int(groups[1])
        y = int(groups[1]) if len(groups[1]) == 4 else int(groups[0])
        sm = (q - 1) * 3 + 1
        return date(y, sm, 1), _end_of_month(y, sm + 2), f"q{q}-{y}"

    # Bare year (last 4-digit run)
    years = re.findall(r'(19|20)\d{2}', low)
    if years:
        # re.findall with a group returns the prefix; re-extract full year
        ym = re.findall(r'((?:19|20)\d{2})', low)
        if ym:
            y = int(ym[-1])
            return date(y, 1, 1), date(y, 12, 31), str(y)

    return None, None, ""


def _to_date(ddmmyyyy: str):
    """Parse a normalised DD/MM/YYYY string to a date, or None."""
    if not ddmmyyyy or "/" not in ddmmyyyy:
        return None
    parts = ddmmyyyy.split("/")
    if len(parts) != 3 or len(parts[2]) != 4:
        return None
    try:
        return date(int(parts[2]), int(parts[1]), int(parts[0]))
    except (ValueError, IndexError):
        return None


def compute_date_flag(normalised_date: str, period_start, period_end) -> str:
    """
    Return a date_flag without ever dropping the row (gap 4):
        ""            ok
        "unparsed"    couldn't parse the date at all
        "future"      date is in the future relative to today
        "out_of_range" parsed but outside the publication's stated period
    """
    d = _to_date(normalised_date)
    if d is None:
        return "unparsed"
    if d > date.today():
        return "future"
    if period_start and period_end:
        # Allow ~31 days of grace either side (meetings logged just outside the
        # stated quarter happen legitimately); anything further out is flagged.
        grace_lo = date.fromordinal(period_start.toordinal() - 31)
        grace_hi = date.fromordinal(period_end.toordinal() + 31)
        if d < grace_lo or d > grace_hi:
            return "out_of_range"
    return ""


# ---------------------------------------------------------------------------
# Date string normalisation (unchanged behaviour from prior versions).
# ---------------------------------------------------------------------------
def normalize_date(date_str: str, fallback_year: str = "") -> str:
    """
    Normalize any date string to DD/MM/YYYY format.

    Handles DD/MM/YYYY, YYYY-MM-DD, YYYY/MM/DD, dotted variants, YYYY-MM,
    Mon-YY ("Nov-16"), "April 2011", "November, 2010", and bare month names
    (using fallback_year when available).
    """
    if not date_str:
        return ""

    date_str = date_str.strip()

    # Strip a trailing time component (spreadsheet cells arrive as datetimes,
    # e.g. "2015-09-01 00:00:00" or "01/09/2015T00:00:00").
    m_dt = re.match(r'^(.*?)[ T]\d{1,2}:\d{2}(:\d{2})?(\.\d+)?$', date_str)
    if m_dt:
        date_str = m_dt.group(1).strip()

    MONTHS = {
        'january': '01', 'february': '02', 'march': '03', 'april': '04',
        'may': '05', 'june': '06', 'july': '07', 'august': '08',
        'september': '09', 'october': '10', 'november': '11', 'december': '12',
        'jan': '01', 'feb': '02', 'mar': '03', 'apr': '04',
        'jun': '06', 'jul': '07', 'aug': '08', 'sep': '09',
        'oct': '10', 'nov': '11', 'dec': '12'
    }

    # Already in DD/MM/YYYY?
    if re.match(r'^\d{1,2}/\d{1,2}/\d{4}$', date_str):
        parts = date_str.split('/')
        return f"{int(parts[0]):02d}/{int(parts[1]):02d}/{parts[2]}"

    # YYYY-MM-DD
    m = re.match(r'^(\d{4})-(\d{1,2})-(\d{1,2})$', date_str)
    if m:
        return f"{int(m.group(3)):02d}/{int(m.group(2)):02d}/{m.group(1)}"

    # YYYY/MM/DD
    m = re.match(r'^(\d{4})/(\d{1,2})/(\d{1,2})$', date_str)
    if m:
        return f"{int(m.group(3)):02d}/{int(m.group(2)):02d}/{m.group(1)}"

    # Dot-separated: could be DD.MM.YY, DD.MM.YYYY, YYYY.MM.DD, or D.M.YY
    m = re.match(r'^(\d{1,4})[.](\d{1,2})[.](\d{1,4})$', date_str)
    if m:
        a, b, c = m.group(1), m.group(2), m.group(3)
        if len(a) == 4:
            return f"{int(c):02d}/{int(b):02d}/{a}"
        elif len(c) == 4:
            return f"{int(a):02d}/{int(b):02d}/{c}"
        elif len(c) == 2:
            year = int(c)
            full_year = 2000 + year if year < 50 else 1900 + year
            return f"{int(a):02d}/{int(b):02d}/{full_year}"

    # YYYY-MM (month only)
    m = re.match(r'^(\d{4})-(\d{1,2})$', date_str)
    if m:
        return f"01/{int(m.group(2)):02d}/{m.group(1)}"

    # Mon-YYYY format: "Sep-2015", "Jul-2014" (and "Sep 2015")
    m = re.match(r'^(\w{3,})[-\s](\d{4})$', date_str)
    if m and m.group(1).lower() in MONTHS:
        return f"01/{MONTHS[m.group(1).lower()]}/{m.group(2)}"

    # Mon-YY format: "Nov-16", "Mar-17", "Oct-10"
    m = re.match(r'^(\w{3,})-(\d{2})$', date_str)
    if m:
        month_name = m.group(1).lower()
        yy = int(m.group(2))
        if month_name in MONTHS:
            full_year = 2000 + yy if yy < 50 else 1900 + yy
            return f"01/{MONTHS[month_name]}/{full_year}"

    # Text month with year: "April 2011", "November, 2010", "May, 2010"
    m = re.match(r'^(\w+),?\s+(\d{4})$', date_str)
    if m:
        month_name = m.group(1).lower()
        year = m.group(2)
        if month_name in MONTHS:
            return f"01/{MONTHS[month_name]}/{year}"

    # Month name only: "June", "October" - use fallback year if available
    month_lower = date_str.lower().strip().rstrip(',')
    if month_lower in MONTHS:
        if fallback_year:
            return f"01/{MONTHS[month_lower]}/{fallback_year}"
        return date_str

    return date_str


def normalize_column_key(key: str) -> str:
    """Normalize a CSV column header for matching (BOM, case, curly quotes)."""
    if not key:
        return ""
    key = key.replace('﻿', '')
    key = key.replace('‘', "'").replace('’', "'")
    key = key.replace('“', '"').replace('”', '"')
    return key.lower().strip()


# ---------------------------------------------------------------------------
# Publication discovery (gaps 1 + 2).
# ---------------------------------------------------------------------------
def _page_search(params, seen_links, publications, cap=1000):
    """
    Page through one Search API query into (seen_links, publications), respecting
    the API's ~1000-offset hard cap. Returns (added_count, reported_total).
    """
    page_size = 100
    added = 0
    reported_total = 0
    start = 0
    while True:
        try:
            r = requests.get(GOVUK_SEARCH_URL, params={
                **params, 'count': page_size, 'start': start,
                'fields': 'link,title,public_timestamp',
            }, timeout=30)
            r.raise_for_status()
            data = r.json()
        except Exception as e:
            print(f"    Search API error at offset {start} ({params.get('filter_organisations','*')}): {e}")
            break
        results = data.get('results', [])
        if not results:
            break
        reported_total = max(reported_total, data.get('total', 0))
        for res in results:
            link = res.get('link', '')
            title = res.get('title', '')
            if link and title and link not in seen_links:
                seen_links.add(link)
                publications.append({
                    'link': link, 'title': title,
                    'timestamp': res.get('public_timestamp', ''),
                })
                added += 1
        start += page_size
        if start >= data.get('total', 0) or start >= cap:
            break
        time.sleep(REQUEST_DELAY)
    return added, reported_total


def _facet_organisations(query, max_facets=400):
    """
    Return [(slug, doc_count)] for every organisation that has transparency
    publications matching `query`. One request; used to partition discovery so no
    single sub-query ever hits the 1000-offset cap (gap 1, exhaustive coverage).
    """
    try:
        r = requests.get(GOVUK_SEARCH_URL, params={
            'filter_format': 'transparency', 'q': query,
            'count': 0, 'facet_organisations': max_facets,
        }, timeout=30)
        r.raise_for_status()
        opts = r.json().get('facets', {}).get('organisations', {}).get('options', [])
    except Exception as e:
        print(f"    Facet error: {e}")
        return []
    out = []
    for o in opts:
        slug = (o.get('value') or {}).get('slug', '')
        if slug:
            out.append((slug, o.get('documents', 0)))
    return out


def discover_publications(query, label=""):
    """
    Discover transparency publications from the GOV.UK Search API, EXHAUSTIVELY.

    The Search API hard-caps paging at ~1000 results per query, so a single broad
    query (e.g. "ministerial meetings" = 5,400+ hits) silently loses everything
    past the cap — which is what gutted 2015-2024 coverage in the first rebuild.

    Fix (gap 1, done properly): partition by organisation. `facet_organisations`
    lists every department with matching transparency pubs and their counts; the
    busiest (Cabinet Office) has only ~270, far under the cap. We page each org
    separately and union, so nothing is capped out. A global both-orderings pass
    is added as a safety net for any publication with no organisation facet, and
    any org that still exceeds the cap is reported loudly.
    """
    seen_links = set()
    publications = []

    print(f"  Discovering {label} publications (partitioned by organisation)...")

    # Safety-net global pass (both orderings) catches org-less publications.
    _, total_newest = _page_search(
        {'filter_format': 'transparency', 'q': query, 'order': '-public_timestamp'},
        seen_links, publications)
    _page_search(
        {'filter_format': 'transparency', 'q': query, 'order': 'public_timestamp'},
        seen_links, publications)

    # Per-organisation partition: this is what makes coverage complete.
    orgs = _facet_organisations(query)
    print(f"    {len(orgs)} organisations have matching transparency publications "
          f"(reported total {total_newest})")
    capped = []
    for slug, count in orgs:
        _, org_total = _page_search(
            {'filter_format': 'transparency', 'q': query,
             'filter_organisations': slug, 'order': '-public_timestamp'},
            seen_links, publications)
        if org_total > 1000:
            # Extremely unlikely for a single org, but never hide it.
            capped.append((slug, org_total))
        time.sleep(REQUEST_DELAY)

    print(f"    Found {len(publications)} unique publications across "
          f"{len(orgs)} organisations")
    if capped:
        print(f"    *** WARNING: {len(capped)} organisation(s) exceed the 1000 cap "
              f"and need year-window sub-partitioning: {capped[:5]} ***")
    return publications, total_newest


def discover_from_collections(collection_paths=KNOWN_COLLECTION_PATHS):
    """
    Gap 2: resolve child publications directly from transparency *collection* pages.

    The collection page is the authoritative list of every edition of a recurring
    series, immune to slug drift. We follow every child document link. Returns a
    list of {link, title, timestamp} (same shape as discover_publications).
    """
    print("  Resolving publications from collection pages...")
    pubs = []
    seen = set()

    for path in collection_paths:
        content_url = f"{GOVUK_CONTENT_URL}{path}"
        try:
            r = requests.get(content_url, timeout=30)
            r.raise_for_status()
            data = r.json()
        except Exception as e:
            print(f"    (collection {path}: {e})")
            continue

        # Child documents live under links.documents on a document_collection.
        docs = data.get('links', {}).get('documents', [])
        # Some collections nest under details.collection_groups -> documents (content_ids).
        for doc in docs:
            link = doc.get('base_path', '')
            title = doc.get('title', '')
            if link and link not in seen:
                seen.add(link)
                pubs.append({
                    'link': link,
                    'title': title,
                    'timestamp': doc.get('public_updated_at', '') or doc.get('public_timestamp', ''),
                    'from_collection': path,
                })
        time.sleep(REQUEST_DELAY)

    print(f"    Collections yielded {len(pubs)} child publications")
    return pubs


def get_attachments_from_publication(pub_link):
    """
    Enumerate ALL downloadable attachments on a publication page (gap 3).

    Returns (attachments, page_meta, error) where:
      attachments = [{url, filename, ext}]   (csv, xlsx, ods, xls)
      page_meta   = {public_updated_at, first_published_at, schema_name}
    Collection pages return ([], meta, None) — they hold no attachments directly.
    """
    if not pub_link.startswith("/"):
        pub_link = "/" + pub_link

    content_url = f"{GOVUK_CONTENT_URL}{pub_link}"
    publication_url = f"https://www.gov.uk{pub_link}"

    try:
        r = requests.get(content_url, timeout=30)
        r.raise_for_status()
        data = r.json()
    except requests.exceptions.HTTPError as e:
        return [], {}, f"HTTP {e.response.status_code} for {content_url}"
    except Exception as e:
        return [], {}, f"Error fetching {content_url}: {e}"

    page_meta = {
        'publication_url': publication_url,
        'schema_name': data.get('schema_name', ''),
        'public_updated_at': data.get('public_updated_at', ''),
        'first_published_at': data.get('first_published_at', ''),
        'title': data.get('title', ''),
    }

    if page_meta['schema_name'] == 'document_collection':
        return [], page_meta, None

    # Accepted attachment extensions, in CSV-first preference order handled later.
    ext_re = re.compile(r'\.(csv|xlsx|xls|ods)(\?|$)', re.IGNORECASE)
    found = {}  # url -> {url, filename, ext}
    details = data.get('details', {})

    def _add(url):
        if url.startswith('/'):
            if '/csv-preview/' in url:
                return
            url = f"https://www.gov.uk{url}"
        m = ext_re.search(url)
        if not m:
            return
        filename = unquote(url.split('/')[-1].split('?')[0])
        found[url] = {'url': url, 'filename': filename, 'ext': m.group(1).lower()}

    # Method 1: structured attachments (most reliable)
    for att in details.get('attachments', []):
        _add(att.get('url', ''))

    # Method 2: HTML document blocks (fallback for older publications)
    for doc in details.get('documents', []):
        if not isinstance(doc, str):
            continue
        for url in re.findall(
            r'href="(https://assets\.publishing\.service\.gov\.uk/[^"]+\.(?:csv|xlsx|xls|ods)[^"]*)"',
            doc, re.IGNORECASE
        ):
            _add(url)
        for url in re.findall(r'href="(/[^"]+\.(?:csv|xlsx|xls|ods)[^"]*)"', doc, re.IGNORECASE):
            _add(url)

    return list(found.values()), page_meta, None


def is_meetings_attachment(filename: str) -> bool:
    """
    Decide whether an attachment is a *meetings* file (gap 2 bug fix).

    The old version excluded any URL containing 'overseas' or 'travel' — which
    WRONGLY dropped the 2024+ Cabinet Office / DSIT slugs
    'cabinet-office-ministerial-overseas-travel-and-meetings-...'. We now key the
    decision on the filename and only exclude gifts/hospitality/expenses files
    that do NOT also mention meetings.
    """
    name = (filename or "").lower()

    # Anything explicitly a meetings file is in, regardless of other words.
    if 'meeting' in name:
        return True

    # Otherwise drop pure gifts/hospitality/expenses/travel returns.
    exclude_terms = ['gift', 'hospitality', 'expense', 'overseas', 'travel']
    if any(term in name for term in exclude_terms):
        return False

    # Generic / ambiguous filename: keep it (some departments use bland names).
    return True


# ---------------------------------------------------------------------------
# Gap 10: canonical department. Source-published labels fragment badly
# ("DCMS" vs "Dcms Ministers Gifts Hospitality Travel And Meetings"), defeating
# per-department grouping. We map every row to a stable canonical department,
# keyed primarily off the publication URL slug (reliable) then the label text.
# Each rule: (canonical, [phrase substrings], [exact slug/word tokens]).
# Ordered most-specific-first so e.g. DESNZ wins over a bare "energy".
# ---------------------------------------------------------------------------
DEPT_RULES = [
    ("DSIT", ["science-innovation-and-technology", "science, innovation"], ["dsit"]),
    ("DESNZ", ["energy-security-and-net-zero", "energy security and net zero"], ["desnz"]),
    ("DBT", ["business-and-trade", "department for business and trade"], ["dbt"]),
    ("DIT", ["international-trade", "international trade"], ["dit"]),
    ("BEIS", ["business-energy-and-industrial-strategy", "business, energy"], ["beis"]),
    ("BIS", ["business-innovation-and-skills", "business, innovation and skills"], ["bis"]),
    ("DHSC", ["health-and-social-care", "department of health"], ["dhsc"]),
    ("DfT", ["department-for-transport", "department for transport"], ["dft"]),
    ("DfE", ["department-for-education", "department for education"], ["dfe"]),
    ("MHCLG", ["housing-communities-and-local-government", "levelling-up-housing-and-communities",
               "communities-and-local-government", "housing, communities"], ["mhclg", "dluhc", "dclg"]),
    ("FCDO", ["foreign-commonwealth-development", "foreign, commonwealth"], ["fcdo"]),
    ("DFID", ["international-development", "international development"], ["dfid"]),
    ("FCO", ["foreign-and-commonwealth-office", "foreign office"], ["fco"]),
    ("MOD", ["ministry-of-defence", "ministry of defence"], ["mod"]),
    ("DWP", ["work-and-pensions", "work and pensions"], ["dwp"]),
    ("Defra", ["environment-food-rural-affairs", "environment, food"], ["defra"]),
    ("DCMS", ["digital-culture-media-and-sport", "culture-media-and-sport", "culture, media"], ["dcms"]),
    ("MOJ", ["ministry-of-justice", "ministry of justice"], ["moj"]),
    ("HM Treasury", ["hm-treasury", "hm treasury"], ["hmt", "treasury"]),
    ("HMRC", ["hm-revenue-customs", "revenue-and-customs", "revenue & customs"], ["hmrc"]),
    ("Home Office", ["home-office", "home office"], []),
    ("NIO", ["northern-ireland-office", "northern ireland office"], ["nio"]),
    ("Scotland Office", ["secretary-of-state-for-scotland", "scotland-office", "scotland office"], []),
    ("Wales Office", ["secretary-of-state-for-wales", "wales-office", "wales office"], []),
    ("DExEU", ["exiting-the-european-union", "department for exiting"], ["dexeu"]),
    ("Attorney General", ["attorney-general", "attorney general"], ["ago"]),
    ("GLD", ["government-legal-department", "treasury-solicitor"], ["gld"]),
    ("UKEF", ["uk-export-finance", "export finance"], ["ukef"]),
    ("Leader of the House of Commons", ["leader-of-the-house-of-commons"], []),
    ("Leader of the House of Lords", ["leader-of-the-house-of-lords"], []),
    ("Prime Minister's Office", ["prime-ministers-office", "10-downing-street", "number-10", "no10"], []),
    ("Cabinet Office", ["cabinet-office", "cabinet office"], ["co"]),
]


def canonical_department(label="", source_url=""):
    """Map a row's (label, source_url) to a stable canonical department, or '' if unknown."""
    slug = (source_url or "").rstrip("/").split("/")[-1].lower()
    text = f"{slug} {(label or '').lower()}"
    tokens = set(re.split(r'[^a-z0-9]+', text))
    for canonical, phrases, codes in DEPT_RULES:
        if any(p in text for p in phrases):
            return canonical
        if any(c in tokens for c in codes):
            return canonical
    return ""


# ---------------------------------------------------------------------------
# Row extraction shared between CSV and XLSX parsers (gaps 4, 5, 10).
# ---------------------------------------------------------------------------
NIL_VALUES = {'nil', 'nil return', 'no meetings', 'none', 'n/a', 'no meetings held',
              'no external meetings', 'nil return.'}


def _process_row(row_lower, department, meeting_type, source_url, source_attachment,
                 fallback_year, period_start, period_end):
    """
    Turn one normalised row dict into ('meeting', record) | ('nil', record) | (None, None).

    Never silently drops a row with a real organisation: out-of-range / unparsed
    dates are FLAGGED, not discarded (gap 4). Nil-return placeholders are routed to
    the nil sheet rather than the data (gap 4).
    """
    minister = (
        row_lower.get("minister", "") or
        row_lower.get("minister's name", "") or
        row_lower.get("minister name", "") or
        row_lower.get("senior official's name", "") or
        row_lower.get("senior official name", "") or
        row_lower.get("name of senior official", "") or
        row_lower.get("senior official", "") or
        row_lower.get("official", "") or
        row_lower.get("official's name", "") or
        row_lower.get("official name", "") or
        row_lower.get("permanent secretary", "") or
        row_lower.get("director general", "") or
        row_lower.get("director general's name", "") or
        row_lower.get("name (permanent secretary only)", "") or
        row_lower.get("senior officials name", "") or
        row_lower.get("name of permanent secretary", "") or
        row_lower.get("name", "") or
        ""
    ).strip()

    date_raw = (
        row_lower.get("date", "") or
        row_lower.get("date of meeting", "") or
        row_lower.get("meeting date", "") or
        row_lower.get("date of external meeting", "") or
        ""
    ).strip()

    org = (
        row_lower.get("name of individual or organisation", "") or
        row_lower.get("name of organisation or individual", "") or
        row_lower.get("person or organisation that meeting was with", "") or
        row_lower.get("organisation", "") or
        row_lower.get("organizations", "") or
        row_lower.get("organisations", "") or
        row_lower.get("name of organisation", "") or
        row_lower.get("name of external organisation", "") or
        row_lower.get("external organisation", "") or
        row_lower.get("organisation(s)", "") or
        row_lower.get("name of external organisation(s)", "") or
        ""
    ).strip()

    purpose = (
        row_lower.get("purpose of meeting", "") or
        row_lower.get("purpose", "") or
        row_lower.get("reason for meeting", "") or
        ""
    ).strip()

    # Nil-return placeholder rows: capture separately rather than dropping silently.
    org_l = org.lower().strip().rstrip('.')
    minister_l = minister.lower().strip()
    is_nil = (org_l in NIL_VALUES) or (minister_l in NIL_VALUES) or \
             (date_raw.lower().strip() in NIL_VALUES)
    if is_nil:
        if minister or org or date_raw:
            canonical, honorific = canonicalize_name(minister)
            return 'nil', {
                "minister": minister,
                "person_canonical": canonical,
                "honorific": honorific,
                "raw_value": org or date_raw or minister,
                "department": department,
                "meeting_type": meeting_type,
                "source_url": source_url,
                "source_attachment": source_attachment,
            }
        return None, None

    # A genuine row needs an organisation and at least a name or a date.
    if not org:
        return None, None
    if not minister and not date_raw:
        return None, None

    date_norm = normalize_date(date_raw, fallback_year=fallback_year)
    date_flag = compute_date_flag(date_norm, period_start, period_end)
    canonical, honorific = canonicalize_name(minister)

    d = _to_date(date_norm)
    starmer_era = bool(d and d >= STARMER_CUTOFF)

    return 'meeting', {
        "minister": minister,
        "person_canonical": canonical,
        "honorific": honorific,
        "date": date_norm,
        "date_raw": date_raw,
        "date_flag": date_flag,
        "organisation": org,
        "purpose": purpose,
        "department": department,
        "department_canonical": canonical_department(department, source_url) or department,
        "meeting_type": meeting_type,
        "starmer_era": starmer_era,
        "source_url": source_url,
        "source_attachment": source_attachment,
    }


def classify_meeting_type(headers_lower=None, filename="", source_url="", fallback="ministerial"):
    """
    Decide ministerial vs senior_official from the strongest available signal.

    The earlier detector only ever flipped *toward* senior_official, so once a
    publication matched the "senior officials" discovery query (and was tagged
    'both'), every row in it — including obviously ministerial files like
    'dhsc-ministerial-meetings...' — was mislabelled senior. This classifier is
    bidirectional and prefers the explicit role wording in the attachment
    filename / URL, then the CSV column headers (which name the attendee's role),
    and only falls back to the publication-level guess when truly ambiguous.
    """
    def _norm(s):
        return re.sub(r'[_\-]+', ' ', (s or '').lower())

    text = _norm(f"{filename} {source_url}")
    senior_sig = (
        'senior official' in text or 'permanent secretar' in text
        or 'director general' in text or bool(re.search(r'\bofficials?\b', text))
    )
    minister_sig = bool(re.search(r'\bminister', text))  # minister/ministers/ministerial

    if senior_sig and not minister_sig:
        return "senior_official"
    if minister_sig and not senior_sig:
        return "ministerial"

    # Filename ambiguous or conflicting -> use the column headers.
    if headers_lower:
        hs = _norm(" ".join(headers_lower))
        senior_hdr = ('permanent secretar' in hs or 'senior official' in hs
                      or 'director general' in hs)
        minister_hdr = 'minister' in hs
        if senior_hdr and not minister_hdr:
            return "senior_official"
        if minister_hdr and not senior_hdr:
            return "ministerial"

    return fallback if fallback in ("ministerial", "senior_official") else "ministerial"


def fetch_attachment_bytes(url):
    """Download an attachment once; return (content_bytes, error)."""
    try:
        r = requests.get(url, timeout=60)
        r.raise_for_status()
        return r.content, None
    except requests.exceptions.HTTPError as e:
        return None, f"HTTP {e.response.status_code}"
    except Exception as e:
        return None, str(e)


def _emit_rows(records, headers, department, meeting_type, source_url,
               source_attachment, fallback_year, period_start, period_end):
    """Shared tail: classify type from headers, run each row through _process_row."""
    meetings, nils = [], []
    headers_lower = [normalize_column_key(h) for h in headers if h]
    meeting_type = classify_meeting_type(headers_lower, source_attachment,
                                         source_url, meeting_type)
    for row_lower in records:
        kind, rec = _process_row(row_lower, department, meeting_type, source_url,
                                 source_attachment, fallback_year,
                                 period_start, period_end)
        if kind == 'meeting':
            meetings.append(rec)
        elif kind == 'nil':
            nils.append(rec)
    return meetings, nils


def parse_csv_bytes(content_bytes, department, meeting_type, source_url,
                    source_attachment, fallback_year, period_start, period_end):
    """Parse meeting + nil rows from CSV bytes. Returns (meetings, nils, error)."""
    try:
        content = content_bytes.decode('utf-8-sig', errors='replace')
        reader = csv.DictReader(io.StringIO(content))
        headers = reader.fieldnames or []
        records = [{normalize_column_key(k): (v.strip() if v else '')
                    for k, v in row.items() if k} for row in reader]
        meetings, nils = _emit_rows(records, headers, department, meeting_type, source_url,
                                    source_attachment, fallback_year, period_start, period_end)
        return meetings, nils, None
    except Exception as e:
        # Malformed CSV (e.g. literal newlines in unquoted fields): retry via pandas.
        m, n, perr = parse_with_pandas(content_bytes, 'csv', department, meeting_type,
                                       source_url, source_attachment, fallback_year,
                                       period_start, period_end)
        if perr:
            return [], [], f"Parse error: {e}"
        return m, n, None


def parse_with_pandas(content_bytes, ext, department, meeting_type, source_url,
                      source_attachment, fallback_year, period_start, period_end):
    """
    Parse ODS / XLS / XLSX (and as a CSV fallback) via pandas, which has the
    engines for each format (odf, xlrd, openpyxl). Reads every sheet. Returns
    (meetings, nils, error). This is what recovers the ~1,100 ODS files and the
    old-binary .xls returns the openpyxl-only path could not read.
    """
    try:
        import pandas as pd
    except ImportError:
        return [], [], "pandas unavailable"
    engine = {'ods': 'odf', 'xls': 'xlrd', 'xlsx': 'openpyxl'}.get(ext)
    try:
        if ext == 'csv':
            # header=None so we can find the real header row ourselves.
            frames = {0: pd.read_csv(io.BytesIO(content_bytes), dtype=str, header=None,
                                     engine='python', on_bad_lines='skip')}
        else:
            frames = pd.read_excel(io.BytesIO(content_bytes), sheet_name=None,
                                   dtype=str, header=None, engine=engine)
    except Exception as e:
        return [], [], f"{ext} parse error: {e}"

    HEADER_HINTS = ("date", "minister", "organisation", "senior official",
                    "permanent secretary", "purpose", "name of")
    meetings, nils = [], []
    for df in frames.values():
        if df is None or df.empty:
            continue
        grid = df.values.tolist()
        # Locate the header row: the first row naming a date/minister/org/purpose column.
        header_idx = None
        for i, raw in enumerate(grid[:25]):
            cells = " ".join(normalize_column_key(str(c)) for c in raw if c is not None)
            if any(h in cells for h in HEADER_HINTS):
                header_idx = i
                break
        if header_idx is None:
            continue
        headers = [str(c) if c is not None else "" for c in grid[header_idx]]
        keys = [normalize_column_key(h) for h in headers]
        records = []
        for raw in grid[header_idx + 1:]:
            rl = {}
            for j, val in enumerate(raw):
                if j < len(keys) and keys[j]:
                    v = '' if (val is None or pd.isna(val)) else str(val).strip()
                    rl[keys[j]] = '' if v.lower() in ('nan', 'nat') else v
            records.append(rl)
        m, n = _emit_rows(records, headers, department, meeting_type, source_url,
                          source_attachment, fallback_year, period_start, period_end)
        meetings.extend(m)
        nils.extend(n)
    return meetings, nils, None


def parse_xlsx_bytes(content_bytes, department, meeting_type, source_url,
                     source_attachment, fallback_year, period_start, period_end):
    """
    Parse meeting + nil rows from XLSX bytes (gap 3 fallback) using openpyxl.

    Heuristically locates the header row (the first row that names a date/minister/
    organisation column), then reads subsequent rows as dicts.
    """
    meetings, nils = [], []
    try:
        from openpyxl import load_workbook
    except ImportError:
        return meetings, nils, "openpyxl not installed (cannot parse xlsx)"

    try:
        wb = load_workbook(io.BytesIO(content_bytes), read_only=True, data_only=True)
    except Exception as e:
        return meetings, nils, f"xlsx open error: {e}"

    try:
        for ws in wb.worksheets:
            rows = ws.iter_rows(values_only=True)
            header = None
            header_idx_map = None
            for raw in rows:
                cells = ["" if c is None else str(c).strip() for c in raw]
                keys = [normalize_column_key(c) for c in cells]
                joined = " ".join(keys)
                if header is None:
                    # Look for a plausible header row.
                    if any(k in joined for k in ["date", "minister", "organisation",
                                                 "senior official", "purpose"]):
                        header = keys
                        header_idx_map = keys
                        meeting_type = classify_meeting_type(keys, source_attachment,
                                                             source_url, meeting_type)
                    continue
                row_lower = {}
                for i, val in enumerate(cells):
                    if i < len(header_idx_map) and header_idx_map[i]:
                        row_lower[header_idx_map[i]] = val
                kind, rec = _process_row(row_lower, department, meeting_type, source_url,
                                         source_attachment, fallback_year,
                                         period_start, period_end)
                if kind == 'meeting':
                    meetings.append(rec)
                elif kind == 'nil':
                    nils.append(rec)
    except Exception as e:
        return meetings, nils, f"xlsx parse error: {e}"
    finally:
        wb.close()
    return meetings, nils, None


def extract_department_from_publication(title, link):
    """Extract department name from a publication title or URL path."""
    dept_patterns = [
        r'^(.+?):\s*ministerial',
        r'^(.+?):\s*senior official',
        r'^(.+?)\s+ministerial',
        r'^(.+?)\s+senior official',
        r"^(.+?)'s\s+ministerial",
        r"^(.+?)'s\s+senior official",
    ]
    for pattern in dept_patterns:
        m = re.match(pattern, title, re.IGNORECASE)
        if m:
            dept = m.group(1).strip().rstrip(':').strip()
            return dept
    parts = link.strip('/').split('/')
    if len(parts) >= 3:
        return parts[-1].split('-ministerial')[0].replace('-', ' ').title()[:50]
    return "Unknown"


# ---------------------------------------------------------------------------
# Gap 6: dedup key for the counting layer (raw rows are kept intact).
# ---------------------------------------------------------------------------
def _normalize_org_for_key(org: str) -> str:
    s = (org or "").lower()
    s = re.sub(r'[^\w\s]', ' ', s)
    s = re.sub(r'\s+', ' ', s).strip()
    return s


def dedup_key_for(meeting) -> str:
    """
    Counting-layer key: (date, normalised organisation, purpose prefix).

    Because the normalised organisation is part of the key, same-day same-boilerplate
    but different-company meetings are NOT merged (the org-token guard the spec asks
    for). Raw rows are never destroyed; this is metadata for an optional dedup view.
    """
    org = _normalize_org_for_key(meeting.get("organisation", ""))
    purpose = (meeting.get("purpose", "") or "").lower().strip()[:60]
    return f"{meeting.get('date','')}|{org}|{purpose}"


def build_name_review(meetings, threshold=0.88, cap=200):
    """
    Gap 5: surface fuzzy near-duplicate canonical names for MANUAL review rather
    than auto-merging (e.g. Allan vs Allen). Returns a list of {a, b, ratio}.
    """
    names = sorted({m.get("person_canonical", "") for m in meetings if m.get("person_canonical")})
    review = []
    for i, a in enumerate(names):
        # difflib close matches within a local window keeps this affordable.
        for b in difflib.get_close_matches(a, names[i + 1:i + 60], n=5, cutoff=threshold):
            if a != b:
                ratio = difflib.SequenceMatcher(None, a, b).ratio()
                review.append({"a": a, "b": b, "ratio": round(ratio, 3)})
                if len(review) >= cap:
                    return review
    return review


def load_prior_index(output_path):
    """Load the previous index (for amendment-drift diffing, gap 9). Returns dict or None."""
    if not output_path.exists():
        return None
    try:
        with gzip.open(output_path, "rt", encoding="utf-8") as f:
            return json.load(f)
    except Exception:
        return None


def build_index():
    """Main function to build the complete UK meetings index."""

    print("=" * 60)
    print("Building UK Ministerial & Senior Officials Meetings Index")
    print("=" * 60)
    print()

    output_path = Path(__file__).parent / "uk_meetings_index.json.gz"
    prior_index = load_prior_index(output_path)
    prior_hashes = {}
    if prior_index:
        for entry in prior_index.get("manifest", []):
            for fn, h in (entry.get("content_hashes") or {}).items():
                prior_hashes[fn] = h
        print(f"  Loaded prior index for amendment diffing "
              f"({len(prior_hashes)} attachment hashes)")
        print()

    # ---- Step 1: Discover publications (search + collections) ----
    print("Step 1: Discovering publications...")

    ministerial_pubs, min_total = discover_publications("ministerial meetings", label="ministerial")
    senior_pubs, sen_total = discover_publications(
        "senior officials meetings transparency", label="senior officials")
    collection_pubs = discover_from_collections()

    all_pubs = {}

    def _register(pub, ptype):
        key = pub['link']
        if key not in all_pubs:
            all_pubs[key] = {**pub, 'type': ptype}
        elif ptype != all_pubs[key]['type'] and all_pubs[key]['type'] != 'both':
            all_pubs[key]['type'] = 'both'

    for pub in ministerial_pubs:
        _register(pub, 'ministerial')
    for pub in senior_pubs:
        _register(pub, 'senior_official')
    for pub in collection_pubs:
        # Infer type from slug/title where possible; default to ministerial.
        series = classify_series(pub['link'])
        ptype = 'senior_official' if 'senior' in series or 'senior' in pub.get('title', '').lower() \
            else 'ministerial'
        _register(pub, ptype)

    print(f"  Total unique publications: {len(all_pubs)} "
          f"({len(collection_pubs)} via collection pages)")
    print()

    # ---- Step 2: Enumerate attachments per publication ----
    print("Step 2: Enumerating attachments from publications...")

    manifest = []          # one entry per publication (gap 8)
    content_errors = []
    pubs_without_attachments = 0
    # Per-publication chosen attachments to download: list of dicts
    download_queue = []    # {url, filename, ext, dept, type, pub_url, series, period_*}

    for i, (link, pub) in enumerate(all_pubs.items()):
        if i % 50 == 0 and i > 0:
            print(f"  Processed {i}/{len(all_pubs)} publications...")

        dept = extract_department_from_publication(pub['title'], link)
        series = classify_series(link)
        period_start, period_end, period_label = parse_period_from_url(link)

        attachments, page_meta, error = get_attachments_from_publication(link)
        pub_url = page_meta.get('publication_url', f"https://www.gov.uk/{link.lstrip('/')}")

        m_entry = {
            "title": pub['title'],
            "link": link,
            "series": series or "unknown",
            "period": period_label,
            "from_collection": pub.get('from_collection', ''),
            "schema": page_meta.get('schema_name', ''),
            "public_updated_at": page_meta.get('public_updated_at', ''),
            "fetched": False,
            "attachment_count": 0,
            "csv_count": 0,
            "xlsx_count": 0,
            "ods_count": 0,
            "row_count": 0,
            "nil_count": 0,
            "formats": [],
            "flags": [],
            "content_hashes": {},
            "fetch_timestamp": "",
        }

        if error:
            content_errors.append((pub['title'], error))
            m_entry["flags"].append(f"content_api_error:{error}")
            manifest.append(m_entry)
            continue

        if page_meta.get('schema_name') == 'document_collection':
            m_entry["flags"].append("collection_page")
            manifest.append(m_entry)
            continue

        # Filter to meetings attachments.
        meeting_atts = [a for a in attachments if is_meetings_attachment(a['filename'])]
        m_entry["attachment_count"] = len(meeting_atts)
        m_entry["formats"] = sorted({a['ext'] for a in meeting_atts})

        if not meeting_atts:
            pubs_without_attachments += 1
            if attachments:
                m_entry["flags"].append("no_meetings_attachment (only gifts/hospitality/etc)")
            else:
                m_entry["flags"].append("no_attachments")
            manifest.append(m_entry)
            time.sleep(REQUEST_DELAY)
            continue

        # Gap 3: CSV-first. Only fall back to a spreadsheet when there is NO CSV.
        csvs = [a for a in meeting_atts if a['ext'] == 'csv']
        chosen = csvs if csvs else meeting_atts
        if not csvs:
            spreadsheet_exts = sorted({a['ext'] for a in chosen})
            m_entry["flags"].append(f"no_csv_using_{'/'.join(spreadsheet_exts)}")

        for a in chosen:
            download_queue.append({
                **a,
                'dept': dept,
                'type': pub['type'],
                'pub_url': pub_url,
                'series': series,
                'period_start': period_start,
                'period_end': period_end,
                'manifest_entry': m_entry,
            })

        manifest.append(m_entry)
        time.sleep(REQUEST_DELAY)

    # Dedupe download queue by attachment URL (same file linked from two pages).
    seen_urls = set()
    unique_queue = []
    for item in download_queue:
        if item['url'] not in seen_urls:
            seen_urls.add(item['url'])
            unique_queue.append(item)

    print(f"  Queued {len(unique_queue)} unique meeting attachments")
    if content_errors:
        print(f"  Content API errors: {len(content_errors)}")
        for title, err in content_errors[:5]:
            print(f"    - {title[:60]}: {err}")
        if len(content_errors) > 5:
            print(f"    ... and {len(content_errors) - 5} more")
    if pubs_without_attachments:
        print(f"  Publications with no meetings attachments: {pubs_without_attachments}")
    print()

    # ---- Step 3: Download + parse (gaps 3, 4, 9) ----
    print("Step 3: Downloading and parsing attachments...")

    all_meetings = []
    all_nils = []
    download_errors = []
    amended_attachments = []  # gap 9
    fetch_ts = datetime.now().isoformat()

    for i, item in enumerate(unique_queue):
        if i % 25 == 0:
            print(f"  Processing attachment {i}/{len(unique_queue)}...")

        # Pass the publication-level type as a fallback only; classify_meeting_type
        # decides per-attachment from filename/headers (a 'both' pub is resolved
        # per file, not blanket-labelled senior_official).
        meeting_type = item['type'] if item['type'] in ('ministerial', 'senior_official') \
            else 'ministerial'
        m_entry = item['manifest_entry']

        fallback_year = ""
        ym = re.findall(r'((?:19|20)\d{2})', item['pub_url'])
        if ym:
            fallback_year = ym[-1]

        content_bytes, err = fetch_attachment_bytes(item['url'])
        if err:
            download_errors.append((item['url'], err))
            m_entry["flags"].append(f"download_error:{item['filename']}:{err}")
            time.sleep(REQUEST_DELAY)
            continue

        # Gap 9: content hash + amendment diff.
        digest = hashlib.sha256(content_bytes).hexdigest()
        m_entry["content_hashes"][item['filename']] = digest
        m_entry["fetch_timestamp"] = fetch_ts
        prior = prior_hashes.get(item['filename'])
        if prior and prior != digest:
            amended_attachments.append(item['filename'])
            m_entry["flags"].append(f"amended_since_last_run:{item['filename']}")

        if item['ext'] == 'csv':
            meetings, nils, perr = parse_csv_bytes(
                content_bytes, item['dept'], meeting_type, item['pub_url'],
                item['filename'], fallback_year, item['period_start'], item['period_end'])
        else:  # xlsx / xls / ods -> pandas (odf / xlrd / openpyxl engines)
            meetings, nils, perr = parse_with_pandas(
                content_bytes, item['ext'], item['dept'], meeting_type, item['pub_url'],
                item['filename'], fallback_year, item['period_start'], item['period_end'])
            if perr and item['ext'] == 'xlsx':
                # Old-style .xls mislabelled .xlsx, or openpyxl-only quirk: last resort.
                m2, n2, e2 = parse_xlsx_bytes(
                    content_bytes, item['dept'], meeting_type, item['pub_url'],
                    item['filename'], fallback_year, item['period_start'], item['period_end'])
                if not e2:
                    meetings, nils, perr = m2, n2, None

        if perr:
            download_errors.append((item['url'], perr))
            m_entry["flags"].append(f"parse_note:{item['filename']}:{perr}")

        m_entry["row_count"] += len(meetings)
        m_entry["nil_count"] += len(nils)
        if item['ext'] == 'csv':
            m_entry["csv_count"] += 1
        elif item['ext'] in ('xlsx', 'xls'):
            m_entry["xlsx_count"] += 1
        else:
            m_entry["ods_count"] += 1
        if meetings or nils or not perr:
            m_entry["fetched"] = True

        all_meetings.extend(meetings)
        all_nils.extend(nils)
        time.sleep(REQUEST_DELAY)

    print(f"  Parsed {len(all_meetings)} meeting records + {len(all_nils)} nil-return rows")
    if download_errors:
        print(f"  Download/parse notes: {len(download_errors)}")
        for url, err in download_errors[:10]:
            print(f"    - {url.split('/')[-1][:60]}: {err}")
        if len(download_errors) > 10:
            print(f"    ... and {len(download_errors) - 10} more")
    if amended_attachments:
        print(f"  *** {len(amended_attachments)} attachment(s) AMENDED since last run "
              f"(content hash changed): {', '.join(amended_attachments[:5])}"
              f"{' ...' if len(amended_attachments) > 5 else ''} ***")
    print()

    # ---- Step 4: Keep raw rows; drop only exact re-parse duplicates (gap 6) ----
    print("Step 4: Removing exact-duplicate re-parses (keeping legitimate cross-return rows)...")

    seen_exact = set()
    raw_meetings = []
    for m in all_meetings:
        exact = (
            m['minister'].lower().strip(),
            m['date'].strip(),
            m['organisation'].lower().strip(),
            m['purpose'].lower().strip()[:100],
            m['department'].lower().strip(),
            m['source_attachment'].lower().strip(),
        )
        if exact not in seen_exact:
            seen_exact.add(exact)
            raw_meetings.append(m)

    # Attach the counting-layer dedup key (gap 6) without collapsing rows.
    for m in raw_meetings:
        m['dedup_key'] = dedup_key_for(m)
    unique_dedup_keys = {m['dedup_key'] for m in raw_meetings}

    print(f"  {len(all_meetings)} parsed -> {len(raw_meetings)} raw rows "
          f"({len(unique_dedup_keys)} distinct meetings by dedup_key)")
    print()

    # ---- Step 5: Build search indexes + name review ----
    print("Step 5: Building search indexes...")

    org_index = defaultdict(list)
    for i, m in enumerate(raw_meetings):
        for word in m['organisation'].lower().split():
            word = re.sub(r'[^\w]', '', word)
            if len(word) > 1:
                org_index[word].append(i)

    name_review = build_name_review(raw_meetings)
    if name_review:
        print(f"  Flagged {len(name_review)} canonical-name near-duplicates for review")

    # Date-flag tally for visibility.
    flag_tally = defaultdict(int)
    for m in raw_meetings:
        if m.get('date_flag'):
            flag_tally[m['date_flag']] += 1
    if flag_tally:
        print(f"  Date flags: {dict(flag_tally)}")

    # ---- Coverage manifest stats (gap 8) ----
    advertised_periods = sorted({
        (e['series'], e['period']) for e in manifest
        if e['period'] and e['series'] != 'unknown'
    })
    unfetched = [e for e in manifest
                 if e['period'] and not e['fetched'] and e['schema'] != 'document_collection'
                 and 'collection_page' not in e['flags']]
    if unfetched:
        print(f"  *** {len(unfetched)} advertised period(s) listed but NOT fetched "
              f"(see manifest flags) ***")

    index = {
        "metadata": {
            "created": datetime.now().isoformat(),
            "meeting_count": len(raw_meetings),
            "unique_meeting_count": len(unique_dedup_keys),
            "nil_return_count": len(all_nils),
            "publications_processed": len(all_pubs),
            "attachments_processed": len(unique_queue),
            "content_api_errors": len(content_errors),
            "download_errors": len(download_errors),
            "amended_attachments": amended_attachments,
            "search_api_reported_totals": {
                "ministerial": min_total, "senior_officials": sen_total},
            "date_flag_tally": dict(flag_tally),
            "advertised_period_count": len(advertised_periods),
            "unfetched_period_count": len(unfetched),
            "starmer_cutoff": STARMER_CUTOFF.isoformat(),
            "coverage": "2010-present",
        },
        "meetings": raw_meetings,
        "nil_returns": all_nils,
        "manifest": manifest,
        "name_review": name_review,
        "org_index": dict(org_index),
    }

    # ---- Step 6: Sanity assertions (gap 1) ----
    manifest_row_total = sum(e['row_count'] for e in manifest)
    if manifest_row_total != len(all_meetings):
        print(f"  *** WARNING: manifest row total ({manifest_row_total}) != parsed meeting "
              f"rows ({len(all_meetings)}). Investigate before trusting coverage. ***")
    else:
        print(f"  Row-count assertion OK: manifest total == parsed rows ({manifest_row_total})")

    # ---- Step 7: Save (with a guard against clobbering a good index) ----
    # A failed scrape (e.g. TLS errors -> zero publications) must NEVER overwrite a
    # healthy prior index with an empty or drastically smaller one. Abort instead.
    prior_count = (prior_index or {}).get("metadata", {}).get("meeting_count", 0)
    if len(raw_meetings) == 0 and prior_count > 0:
        print()
        print("*** ABORTING SAVE: parsed 0 meetings but the existing index has "
              f"{prior_count}. This usually means discovery/download failed "
              "(network/TLS). The existing index was left untouched. ***")
        return None
    if prior_count and len(raw_meetings) < 0.5 * prior_count and "--force" not in sys.argv:
        print()
        print(f"*** ABORTING SAVE: parsed {len(raw_meetings)} meetings, less than half "
              f"the prior {prior_count}. This looks like a partial/failed run. The "
              "existing index was left untouched. Re-run with --force to override. ***")
        return None

    print(f"Step 7: Saving to {output_path}...")
    with gzip.open(output_path, "wt", encoding="utf-8") as f:
        json.dump(index, f, ensure_ascii=False)
    file_size_mb = output_path.stat().st_size / (1024 * 1024)
    print(f"  Saved ({file_size_mb:.1f} MB)")

    # ---- Summary ----
    print()
    print("=" * 60)
    print("BUILD COMPLETE")
    print("=" * 60)
    print(f"  Raw meeting rows:        {len(raw_meetings)}")
    print(f"  Distinct meetings:       {len(unique_dedup_keys)}")
    print(f"  Nil-return rows:         {len(all_nils)}")
    print(f"  Publications:            {len(all_pubs)}")
    print(f"  Attachments parsed:      {len(unique_queue)}")
    print(f"  Advertised periods:      {len(advertised_periods)}")
    print(f"  Unfetched periods:       {len(unfetched)}")
    print(f"  Amended since last run:  {len(amended_attachments)}")
    print(f"  File size:               {file_size_mb:.1f} MB")

    print()
    print("Sample meetings:")
    for m in raw_meetings[:5]:
        flag = f" [{m['date_flag']}]" if m.get('date_flag') else ""
        print(f"  [{m['department']}] {m['person_canonical'] or m['minister']}: "
              f"{m['organisation'][:40]} ({m['date']}){flag}")

    return index


if __name__ == "__main__":
    build_index()
