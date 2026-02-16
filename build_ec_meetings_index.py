#!/usr/bin/env python3
"""
Build a pre-indexed database of all European Commission meetings with
interest representatives.

Downloads XLSX files from the EC Open Data Portal, combines them into a
single searchable JSON file, and compresses it.

Data sources (all from ec.europa.eu/transparencyinitiative/meetings/):
1. Von der Leyen Commission II (2024-2029) - Commissioners & Cabinet members
2. Von der Leyen Commission I (2019-2024) - Commissioners & Cabinet members
3. Directors-General meetings

Run periodically (e.g. weekly via GitHub Actions) to keep up to date.

Usage:
    python build_ec_meetings_index.py

Output:
    ec_meetings_index.json.gz - Complete index of all EC meetings
"""

import json
import gzip
import os
import sys
import re
import requests
import tempfile
from datetime import datetime
from collections import defaultdict

try:
    import openpyxl
except ImportError:
    print("Installing openpyxl...")
    os.system(f"{sys.executable} -m pip install openpyxl --break-system-packages -q")
    import openpyxl


# EC Open Data Portal endpoints
EC_DATASETS = [
    {
        "name": "Von der Leyen II (2024-2029)",
        "url": "https://ec.europa.eu/transparencyinitiative/meetings/dataxlsx.do?name=meetingscommissionrepresentatives2429",
        "type": "commissioner",
        "period": "2024-2029",
    },
    {
        "name": "Von der Leyen I (2019-2024)",
        "url": "https://ec.europa.eu/transparencyinitiative/meetings/dataxlsx.do?name=meetingscommissionrepresentatives1924",
        "type": "commissioner",
        "period": "2019-2024",
    },
    {
        "name": "Directors-General",
        "url": "https://ec.europa.eu/transparencyinitiative/meetings/dataxlsx.do?name=meetingsdirectorgenerals",
        "type": "dg",
        "period": "2014-present",
    },
]

SOURCE_URL = "https://data.europa.eu/data/datasets/european-commission-meetings-with-interest-representatives?locale=en"


def download_xlsx(url: str) -> str:
    """Download an XLSX file to a temp path."""
    print(f"  Downloading {url}...")
    response = requests.get(url, timeout=120)
    response.raise_for_status()
    
    tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
    tmp.write(response.content)
    tmp.close()
    print(f"  Downloaded {len(response.content):,} bytes")
    return tmp.name


def parse_commissioner_xlsx(filepath: str, dataset_info: dict) -> list:
    """
    Parse a Commissioner/Cabinet meetings XLSX.
    
    Expected headers (row 2):
        Name of cabinet | Name of EC representative | Title of EC representative |
        Date of meeting | Location | Name of interest representative |
        Transparency register ID | Subject of the meeting
    """
    meetings = []
    wb = openpyxl.load_workbook(filepath, read_only=True)
    ws = wb.active
    
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    
    # Find header row (usually row 2)
    header_row_idx = None
    for i, row in enumerate(rows[:5]):
        if row and any(str(c or "").lower().startswith("name of") for c in row):
            header_row_idx = i
            break
    
    if header_row_idx is None:
        print(f"  WARNING: Could not find header row")
        return meetings
    
    headers = [str(c or "").strip().lower() for c in rows[header_row_idx]]
    
    for row in rows[header_row_idx + 1:]:
        if not row or not any(row):
            continue
        
        row_dict = {}
        for j, val in enumerate(row):
            if j < len(headers):
                row_dict[headers[j]] = str(val).strip() if val else ""
        
        # Extract fields - Commissioner meetings
        cabinet = row_dict.get("name of cabinet", "")
        representative = row_dict.get("name of ec representative", "")
        title = row_dict.get("title of ec representative", "")
        date = row_dict.get("date of meeting", "")
        location = row_dict.get("location", "")
        org = row_dict.get("name of interest representative", "")
        tr_id = row_dict.get("transparency register id", "")
        subject = row_dict.get("subject of the meeting", "").replace("\r\n", " ").strip()
        
        if not org and not representative:
            continue
        
        # Normalize date to DD/MM/YYYY
        if date and re.match(r"\d{4}-\d{2}-\d{2}", date):
            parts = date.split("-")
            date = f"{parts[2]}/{parts[1]}/{parts[0]}"
        
        meetings.append({
            "cabinet": cabinet,
            "representative": representative,
            "title": title,
            "date": date,
            "location": location,
            "organisation": org,
            "transparency_register_id": tr_id,
            "subject": subject,
            "dataset": dataset_info["name"],
            "period": dataset_info["period"],
            "type": dataset_info["type"],
        })
    
    return meetings


def parse_dg_xlsx(filepath: str, dataset_info: dict) -> list:
    """
    Parse a Directors-General meetings XLSX.
    
    Expected headers (row 2):
        Name of DG/EA - full name | Name of DG/EA - acronym |
        Name of EC/EA representative | Title of EC/EA representative |
        Date of meeting | Location | Name of interest representative |
        Transparency register ID | Subject of the meeting
    """
    meetings = []
    wb = openpyxl.load_workbook(filepath, read_only=True)
    ws = wb.active
    
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    
    # Find header row
    header_row_idx = None
    for i, row in enumerate(rows[:5]):
        if row and any(str(c or "").lower().startswith("name of") for c in row):
            header_row_idx = i
            break
    
    if header_row_idx is None:
        print(f"  WARNING: Could not find header row in DG file")
        return meetings
    
    headers = [str(c or "").strip().lower() for c in rows[header_row_idx]]
    
    for row in rows[header_row_idx + 1:]:
        if not row or not any(row):
            continue
        
        row_dict = {}
        for j, val in enumerate(row):
            if j < len(headers):
                row_dict[headers[j]] = str(val).strip() if val else ""
        
        # Extract fields - DG meetings
        dg_full = row_dict.get("name of dg/ea - full name", "")
        dg_acronym = row_dict.get("name of dg/ea - acronym", "")
        representative = row_dict.get("name of ec/ea representative", 
                        row_dict.get("name of ec representative", ""))
        title = row_dict.get("title of ec/ea representative",
                row_dict.get("title of ec representative", ""))
        date = row_dict.get("date of meeting", "")
        location = row_dict.get("location", "")
        org = row_dict.get("name of interest representative", "")
        tr_id = row_dict.get("transparency register id", "")
        subject = row_dict.get("subject of the meeting", "").replace("\r\n", " ").strip()
        
        if not org and not representative:
            continue
        
        # Build cabinet equivalent for DGs
        cabinet = f"{dg_full} ({dg_acronym})" if dg_acronym else dg_full
        
        # Normalize date
        if date and re.match(r"\d{4}-\d{2}-\d{2}", date):
            parts = date.split("-")
            date = f"{parts[2]}/{parts[1]}/{parts[0]}"
        
        meetings.append({
            "cabinet": cabinet,
            "representative": representative,
            "title": title,
            "date": date,
            "location": location,
            "organisation": org,
            "transparency_register_id": tr_id,
            "subject": subject,
            "dataset": dataset_info["name"],
            "period": dataset_info["period"],
            "type": dataset_info["type"],
        })
    
    return meetings


def build_index():
    """Download all EC meeting data and build the index."""
    
    print("=" * 60)
    print("EC Meetings Index Builder")
    print("=" * 60)
    
    all_meetings = []
    
    for dataset in EC_DATASETS:
        print(f"\n--- {dataset['name']} ---")
        
        try:
            filepath = download_xlsx(dataset["url"])
            
            if dataset["type"] == "dg":
                meetings = parse_dg_xlsx(filepath, dataset)
            else:
                meetings = parse_commissioner_xlsx(filepath, dataset)
            
            print(f"  Parsed {len(meetings)} meetings")
            all_meetings.extend(meetings)
            
            # Clean up
            os.unlink(filepath)
            
        except Exception as e:
            print(f"  ERROR: {e}")
            continue
    
    print(f"\n{'=' * 60}")
    print(f"Total meetings: {len(all_meetings)}")
    
    # Deduplicate (some meetings might appear in multiple datasets)
    seen = set()
    unique = []
    for m in all_meetings:
        key = (m["date"], m["representative"], m["organisation"])
        if key not in seen:
            seen.add(key)
            unique.append(m)
    
    print(f"After dedup: {len(unique)} unique meetings")
    
    # Build metadata
    representatives = set()
    cabinets = set()
    organisations = set()
    
    for m in unique:
        if m["representative"]:
            representatives.add(m["representative"])
        if m["cabinet"]:
            cabinets.add(m["cabinet"])
        if m["organisation"]:
            organisations.add(m["organisation"])
    
    index = {
        "metadata": {
            "created": datetime.now().isoformat(),
            "source": SOURCE_URL,
            "datasets": [d["name"] for d in EC_DATASETS],
            "total_meetings": len(unique),
            "unique_representatives": len(representatives),
            "unique_cabinets": len(cabinets),
            "unique_organisations": len(organisations),
            "coverage": "2014-present",
        },
        "meetings": unique,
    }
    
    # Save compressed
    output_path = "ec_meetings_index.json.gz"
    with gzip.open(output_path, "wt", encoding="utf-8") as f:
        json.dump(index, f)
    
    file_size = os.path.getsize(output_path)
    print(f"\nSaved to {output_path} ({file_size / 1024 / 1024:.1f} MB)")
    print(f"  {len(unique)} meetings")
    print(f"  {len(representatives)} unique EC representatives")
    print(f"  {len(cabinets)} unique cabinets/DGs")
    print(f"  {len(organisations)} unique organisations")
    
    return index


if __name__ == "__main__":
    build_index()
